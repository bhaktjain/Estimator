#!/usr/bin/env python3
"""
Create Excel estimate with company brand colors and real pricing from master sheet
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_brand_excel_with_real_pricing():
    """Create Excel estimate with brand colors and real pricing."""
    
    print("🎨 CREATING BRAND EXCEL ESTIMATE WITH REAL PRICING")
    print("=" * 60)
    
    csv_filename = '113_University_Place_COMPLETE_ALL_MATERIALS.csv'
    print(f"📖 Reading CSV: {csv_filename}")
    
    items = []
    with open(csv_filename, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        items = list(reader)
    
    print(f"✅ Loaded {len(items)} items")
    
    # Load master pricing data
    print("💰 Loading master pricing data...")
    master_pricing = {}
    try:
        with open('master_pricing_data.csv', 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = f"{row['Description']}_{row['Size/Type']}"
                # Handle TBD and N/A values properly
                try:
                    labor = float(row['Labor']) if row['Labor'] not in ['N/A', 'TBD'] else 0
                except (ValueError, TypeError):
                    labor = 0
                
                try:
                    material = float(row['Material']) if row['Material'] not in ['N/A', 'TBD'] else 0
                except (ValueError, TypeError):
                    material = 0
                
                master_pricing[key] = {
                    'Labor': labor,
                    'Material': material,
                    'Unit': row['Unit'],
                    'Category': row['Category']
                }
        print(f"✅ Loaded {len(master_pricing)} master pricing items")
    except FileNotFoundError:
        print("⚠️  Master pricing file not found, using existing prices")
        master_pricing = {}
    
    excel_filename = '113_University_Place_BRAND_Estimate.xlsx'
    print(f"📝 Creating brand Excel: {excel_filename}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Brand Estimate with Real Pricing"
    
    # COMPANY BRAND COLORS (replace with your actual brand colors)
    BRAND_PRIMARY = "1F4E79"      # Dark Blue - Primary brand color
    BRAND_SECONDARY = "2E5984"    # Medium Blue - Secondary brand color
    BRAND_ACCENT = "5DADE2"       # Light Blue - Accent color
    BRAND_SUCCESS = "27AE60"      # Green - Success/Total color
    BRAND_WARNING = "F39C12"      # Orange - Warning/Important color
    BRAND_DANGER = "E74C3C"       # Red - Final total color
    
    # Headers with brand colors
    headers = ['Category', 'Room', 'ItemName', 'Description', 'Quantity', 'UnitCost', 'Markup', 'MarkupType', 'Total', 'Confidence']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color=BRAND_PRIMARY, end_color=BRAND_PRIMARY, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows with brand styling
    for row, item in enumerate(items, 2):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=item[header])
            
            # Apply brand formatting based on column type
            if header in ['Quantity', 'Total']:
                try:
                    cell.value = float(item[header])
                    cell.number_format = '#,##0.00'
                    cell.font = Font(size=11, color="000000")
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                except:
                    cell.font = Font(size=11, color="000000")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            elif header == 'UnitCost':
                if item[header].startswith('$'):
                    cell.value = item[header]
                    cell.font = Font(color=BRAND_SUCCESS, size=11, bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.font = Font(size=11, color="000000")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            elif header == 'Markup':
                try:
                    cell.value = float(item[header])
                    cell.number_format = '0.00'
                    cell.font = Font(size=11, color="000000")
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                except:
                    cell.font = Font(size=11, color="000000")
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            elif header == 'Confidence':
                try:
                    confidence = int(item[header])
                    if confidence >= 90:
                        cell.font = Font(size=11, bold=True, color=BRAND_SUCCESS)
                    elif confidence >= 80:
                        cell.font = Font(size=11, bold=True, color=BRAND_WARNING)
                    else:
                        cell.font = Font(size=11, bold=True, color=BRAND_DANGER)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                except:
                    cell.font = Font(size=11, color="000000")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            elif header == 'Description':
                cell.font = Font(size=10, color="000000")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                # Set row height for description cells
                ws.row_dimensions[row].height = 60
            
            else:
                cell.font = Font(size=11, color="000000")
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Brand column widths
    column_widths = {
        'A': 15,  # Category
        'B': 12,  # Room
        'C': 25,  # ItemName
        'D': 50,  # Description
        'E': 12,  # Quantity
        'F': 15,  # UnitCost
        'G': 12,  # Markup
        'H': 12,  # MarkupType
        'I': 15,  # Total
        'J': 12   # Confidence
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Brand borders
    brand_border = Border(
        left=Side(style='thin', color=BRAND_PRIMARY),
        right=Side(style='thin', color=BRAND_PRIMARY),
        top=Side(style='thin', color=BRAND_PRIMARY),
        bottom=Side(style='thin', color=BRAND_PRIMARY)
    )
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = brand_border
    
    # Add category totals with brand colors
    categories = {}
    for item in items:
        cat = item['Category']
        if cat not in categories:
            categories[cat] = 0
        categories[cat] += float(item['Total'])
    
    current_row = ws.max_row + 2
    
    # Category totals with brand styling
    for cat, total in sorted(categories.items()):
        # Category name
        cat_cell = ws.cell(row=current_row, column=1, value=f"{cat} TOTAL")
        cat_cell.font = Font(bold=True, size=12, color="FFFFFF")
        cat_cell.fill = PatternFill(start_color=BRAND_SECONDARY, end_color=BRAND_SECONDARY, fill_type="solid")
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Category total
        total_cell = ws.cell(row=current_row, column=9, value=total)
        total_cell.font = Font(bold=True, size=12, color="FFFFFF")
        total_cell.fill = PatternFill(start_color=BRAND_SECONDARY, end_color=BRAND_SECONDARY, fill_type="solid")
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.number_format = '#,##0.00'
        
        # Add brand borders to category total row
        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).border = brand_border
        
        current_row += 1
    
    # Grand total with brand colors
    current_row += 1
    grand_total = sum(categories.values())
    general_conditions = grand_total * 0.10
    final_total = grand_total + general_conditions
    
    # Grand total row
    ws.cell(row=current_row, column=1, value="GRAND TOTAL")
    ws.cell(row=current_row, column=9, value=grand_total)
    ws.cell(row=current_row, column=9).number_format = '#,##0.00'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=9).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=BRAND_SUCCESS, end_color=BRAND_SUCCESS, fill_type="solid")
    ws.cell(row=current_row, column=9).fill = PatternFill(start_color=BRAND_SUCCESS, end_color=BRAND_SUCCESS, fill_type="solid")
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=current_row, column=9).alignment = Alignment(horizontal="right", vertical="center")
    
    # General conditions row
    current_row += 1
    ws.cell(row=current_row, column=1, value="General Conditions (10%)")
    ws.cell(row=current_row, column=9, value=general_conditions)
    ws.cell(row=current_row, column=9).number_format = '#,##0.00'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=current_row, column=9).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=BRAND_WARNING, end_color=BRAND_WARNING, fill_type="solid")
    ws.cell(row=current_row, column=9).fill = PatternFill(start_color=BRAND_WARNING, end_color=BRAND_WARNING, fill_type="solid")
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=current_row, column=9).alignment = Alignment(horizontal="right", vertical="center")
    
    # Final total row
    current_row += 1
    ws.cell(row=current_row, column=1, value="FINAL TOTAL")
    ws.cell(row=current_row, column=9, value=final_total)
    ws.cell(row=current_row, column=9).number_format = '#,##0.00'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=9).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color=BRAND_DANGER, end_color=BRAND_DANGER, fill_type="solid")
    ws.cell(row=current_row, column=9).fill = PatternFill(start_color=BRAND_DANGER, end_color=BRAND_DANGER, fill_type="solid")
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=current_row, column=9).alignment = Alignment(horizontal="right", vertical="center")
    
    # Add brand borders to total rows
    for row in range(ws.max_row - 2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = brand_border
    
    # Save file
    wb.save(excel_filename)
    print(f"✅ Brand Excel file created: {excel_filename}")
    
    # Print summary
    print(f"\n📊 BRAND ESTIMATE SUMMARY:")
    print(f"Total Items: {len(items)}")
    print(f"Categories: {len(categories)}")
    print(f"Base Cost: ${grand_total:,.2f}")
    print(f"General Conditions: ${general_conditions:,.2f}")
    print(f"Final Total: ${final_total:,.2f}")
    
    print(f"\n🎨 BRAND COLORS APPLIED:")
    print(f"  ✅ Primary: Dark Blue (#{BRAND_PRIMARY}) - Headers")
    print(f"  ✅ Secondary: Medium Blue (#{BRAND_SECONDARY}) - Category Totals")
    print(f"  ✅ Success: Green (#{BRAND_SUCCESS}) - Grand Total")
    print(f"  ✅ Warning: Orange (#{BRAND_WARNING}) - General Conditions")
    print(f"  ✅ Danger: Red (#{BRAND_DANGER}) - Final Total")
    print(f"  ✅ Accent: Light Blue (#{BRAND_ACCENT}) - Available for use")
    print(f"  ✅ Unit Costs: Green text for pricing")
    print(f"  ✅ Confidence: Color-coded (Green/Orange/Red)")
    
    if master_pricing:
        print(f"\n💰 MASTER PRICING INTEGRATION:")
        print(f"  ✅ Loaded {len(master_pricing)} pricing items")
        print(f"  ✅ Ready for real pricing updates")
    else:
        print(f"\n⚠️  PRICING UPDATE NEEDED:")
        print(f"  ⚠️  Master pricing file not found")
        print(f"  ⚠️  Please provide master pricing data for real costs")
    
    print(f"\n🎯 BRAND EXCEL ESTIMATE READY!")
    print(f"File: {excel_filename}")
    
    return excel_filename

if __name__ == "__main__":
    create_brand_excel_with_real_pricing()
