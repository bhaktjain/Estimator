#!/usr/bin/env python3
"""
Create beautiful Excel estimate from comprehensive materials CSV
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_estimate():
    """Create beautiful Excel estimate from CSV."""
    
    print("📊 CREATING COMPLETE EXCEL ESTIMATE")
    print("=" * 50)
    
    # Read the comprehensive materials CSV
    csv_filename = '113_University_Place_Complete_All_Materials_Estimate.csv'
    print(f"📖 Reading CSV: {csv_filename}")
    
    items = []
    with open(csv_filename, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        items = list(reader)
    
    print(f"✅ Loaded {len(items)} items")
    
    # Create Excel file
    excel_filename = '113_University_Place_Complete_All_Materials_Estimate.xlsx'
    print(f"📝 Creating Excel: {excel_filename}")
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complete Materials Estimate"
    
    # Define headers
    headers = ['Category', 'Room', 'ItemName', 'Description', 'Quantity', 'UnitCost', 'Markup', 'MarkupType', 'Total', 'Confidence']
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row, item in enumerate(items, 2):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=item[header])
            
            # Format numbers
            if header in ['Quantity', 'Total']:
                try:
                    cell.value = float(item[header])
                    cell.number_format = '#,##0.00'
                except:
                    pass
            
            # Format currency
            if header == 'UnitCost':
                if item[header].startswith('$'):
                    cell.value = item[header]
                    cell.font = Font(color="006100")  # Green for costs
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f"{column_letter}{row}"].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        # Set reasonable column widths
        if col == 1:  # Category
            ws.column_dimensions[column_letter].width = 26
        elif col == 2:  # Room
            ws.column_dimensions[column_letter].width = 18
        elif col == 3:  # ItemName
            ws.column_dimensions[column_letter].width = 42
        elif col == 4:  # Description
            ws.column_dimensions[column_letter].width = 60
        else:
            ws.column_dimensions[column_letter].width = min(max_length + 2, 20)
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
    
    # Add category totals
    current_category = None
    category_total = 0
    total_row = len(items) + 3
    
    for row, item in enumerate(items, 2):
        if item['Category'] != current_category:
            if current_category:
                # Write previous category total
                ws.cell(row=total_row, column=1, value=f"{current_category} TOTAL:")
                ws.cell(row=total_row, column=9, value=category_total)
                ws.cell(row=total_row, column=9).number_format = '$#,##0.00'
                ws.cell(row=total_row, column=1).font = Font(bold=True)
                ws.cell(row=total_row, column=9).font = Font(bold=True)
                total_row += 1
            
            current_category = item['Category']
            category_total = 0
        
        try:
            category_total += float(item['Total'])
        except:
            pass
    
    # Write final category total
    if current_category:
        ws.cell(row=total_row, column=1, value=f"{current_category} TOTAL:")
        ws.cell(row=total_row, column=9, value=category_total)
        ws.cell(row=total_row, column=9).number_format = '$#,##0.00'
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=9).font = Font(bold=True)
        total_row += 1
    
    # Add grand total
    grand_total = sum(float(item['Total']) for item in items)
    ws.cell(row=total_row + 1, column=1, value="GRAND TOTAL:")
    ws.cell(row=total_row + 1, column=9, value=grand_total)
    ws.cell(row=total_row + 1, column=9).number_format = '$#,##0.00'
    ws.cell(row=total_row + 1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=total_row + 1, column=9).font = Font(bold=True, size=14)
    
    # Save the file
    wb.save(excel_filename)
    print(f"✅ Excel file created: {excel_filename}")
    
    # Print summary
    print(f"\n📊 COMPLETE ALL MATERIALS ESTIMATE SUMMARY:")
    print(f"Total items: {len(items)}")
    print(f"Base Cost: ${grand_total:,.2f}")
    print(f"General Conditions (10%): ${grand_total * 0.10:,.2f}")
    print(f"Final Total: ${grand_total * 1.10:,.2f}")
    
    # Show categories
    categories = {}
    for item in items:
        cat = item['Category']
        categories[cat] = categories.get(cat, 0) + 1
    
    print(f"\n📋 CATEGORIES INCLUDED:")
    for cat, count in sorted(categories.items()):
        print(f"  {cat}: {count} items")
    
    print(f"\n🎯 Complete ALL materials Excel estimate ready: {excel_filename}")

if __name__ == "__main__":
    create_excel_estimate()
