#!/usr/bin/env python3
"""
Create FINAL Excel estimate from complete materials CSV with proper styling
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_final_excel():
    """Create beautiful Excel estimate from complete materials CSV with proper styling."""
    
    print("📊 CREATING FINAL COMPLETE EXCEL ESTIMATE WITH PROPER STYLING")
    print("=" * 60)
    
    csv_filename = '113_University_Place_COMPLETE_ALL_MATERIALS.csv'
    print(f"📖 Reading CSV: {csv_filename}")
    
    items = []
    with open(csv_filename, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        items = list(reader)
    
    print(f"✅ Loaded {len(items)} items")
    
    excel_filename = '113_University_Place_COMPLETE_ALL_MATERIALS.xlsx'
    print(f"📝 Creating Excel: {excel_filename}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Complete All Materials Estimate"
    
    # Headers with proper styling
    headers = ['Category', 'Room', 'ItemName', 'Description', 'Quantity', 'UnitCost', 'Markup', 'MarkupType', 'Total', 'Confidence']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows with proper styling
    for row, item in enumerate(items, 2):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=item[header])
            
            # Apply proper formatting based on column type
            if header in ['Quantity', 'Total']:
                try:
                    cell.value = float(item[header])
                    cell.number_format = '#,##0.00'
                    cell.font = Font(size=11)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                except:
                    cell.font = Font(size=11)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            elif header == 'UnitCost':
                if item[header].startswith('$'):
                    cell.value = item[header]
                    cell.font = Font(color="006100", size=11, bold=True)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.font = Font(size=11)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            elif header == 'Markup':
                try:
                    cell.value = float(item[header])
                    cell.number_format = '0.00'
                    cell.font = Font(size=11)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                except:
                    cell.font = Font(size=11)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            elif header == 'Confidence':
                try:
                    cell.value = int(item[header])
                    cell.font = Font(size=11, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                except:
                    cell.font = Font(size=11)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            elif header == 'Description':
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                # Set row height for description cells
                ws.row_dimensions[row].height = 60
            
            else:
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths with proper sizing
    column_widths = {
        'A': 15,  # Category
        'B': 12,  # Room
        'C': 25,  # ItemName
        'D': 50,  # Description
        'E': 12,  # Quantity
        'F': 15,  # UnitCost
        'G': 12,  # Markup
        'H': 12,  # MarkupType
        'I': 15,  # Total
        'J': 12   # Confidence
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Add borders with proper styling
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
    
    # Add category totals with proper styling
    categories = {}
    for item in items:
        cat = item['Category']
        if cat not in categories:
            categories[cat] = 0
        categories[cat] += float(item['Total'])
    
    current_row = ws.max_row + 2
    
    # Category totals with proper styling
    for cat, total in sorted(categories.items()):
        # Category name
        cat_cell = ws.cell(row=current_row, column=1, value=f"{cat} TOTAL")
        cat_cell.font = Font(bold=True, size=12, color="FFFFFF")
        cat_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Category total
        total_cell = ws.cell(row=current_row, column=9, value=total)
        total_cell.font = Font(bold=True, size=12, color="FFFFFF")
        total_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        total_cell.alignment = Alignment(horizontal="right", vertical="center")
        total_cell.number_format = '#,##0.00'
        
        # Add borders to category total row
        for col in range(1, len(headers) + 1):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1
    
    # Grand total with proper styling
    current_row += 1
    grand_total = sum(categories.values())
    general_conditions = grand_total * 0.10
    final_total = grand_total + general_conditions
    
    # Grand total row
    ws.cell(row=current_row, column=1, value="GRAND TOTAL")
    ws.cell(row=current_row, column=9, value=grand_total)
    ws.cell(row=current_row, column=9).number_format = '#,##0.00'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=9).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.cell(row=current_row, column=9).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=current_row, column=9).alignment = Alignment(horizontal="right", vertical="center")
    
    # General conditions row
    current_row += 1
    ws.cell(row=current_row, column=1, value="General Conditions (10%)")
    ws.cell(row=current_row, column=9, value=general_conditions)
    ws.cell(row=current_row, column=9).number_format = '#,##0.00'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=current_row, column=9).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.cell(row=current_row, column=9).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=current_row, column=9).alignment = Alignment(horizontal="right", vertical="center")
    
    # Final total row
    current_row += 1
    ws.cell(row=current_row, column=1, value="FINAL TOTAL")
    ws.cell(row=current_row, column=9, value=final_total)
    ws.cell(row=current_row, column=9).number_format = '#,##0.00'
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=9).font = Font(bold=True, size=16, color="FFFFFF")
    ws.cell(row=current_row, column=1).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.cell(row=current_row, column=9).fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=current_row, column=9).alignment = Alignment(horizontal="right", vertical="center")
    
    # Add borders to total rows
    for row in range(ws.max_row - 2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
    
    # Save file
    wb.save(excel_filename)
    print(f"✅ Excel file created with proper styling: {excel_filename}")
    
    # Print summary
    print(f"\n📊 FINAL COMPLETE ESTIMATE SUMMARY:")
    print(f"Total Items: {len(items)}")
    print(f"Categories: {len(categories)}")
    print(f"Base Cost: ${grand_total:,.2f}")
    print(f"General Conditions: ${general_conditions:,.2f}")
    print(f"Final Total: ${final_total:,.2f}")
    
    print(f"\n🎨 STYLING APPLIED:")
    print(f"  ✅ Headers: Blue background (#366092) with white text")
    print(f"  ✅ Category Totals: Light blue background (#4472C4) with white text")
    print(f"  ✅ Grand Total: Green background (#70AD47) with white text")
    print(f"  ✅ Final Total: Red background (#C00000) with white text")
    print(f"  ✅ Unit Costs: Green text (#006100)")
    print(f"  ✅ Proper borders and alignment")
    print(f"  ✅ Dynamic row heights for descriptions")
    
    print(f"\n🎯 COMPLETE ALL MATERIALS EXCEL ESTIMATE READY!")
    print(f"File: {excel_filename}")

if __name__ == "__main__":
    create_final_excel()
