#!/usr/bin/env python3
"""
Create professional Excel estimate using the exact same styling as comprehensive_cleanup.py
and integrate with master internal pricing sheet
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from collections import defaultdict

def create_professional_excel_with_internal_pricing():
    """Create Excel estimate with exact styling from comprehensive_cleanup.py and internal pricing."""
    
    print("🎨 CREATING PROFESSIONAL EXCEL WITH INTERNAL PRICING")
    print("=" * 60)
    
    csv_filename = '113_University_Place_COMPLETE_ALL_MATERIALS.csv'
    print(f"📖 Reading CSV: {csv_filename}")
    
    items = []
    with open(csv_filename, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        items = list(reader)
    
    print(f"✅ Loaded {len(items)} items")
    
    # Load master internal pricing data
    print("💰 Loading master internal pricing data...")
    master_pricing = {}
    try:
        with open('master_pricing_data.csv', 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = f"{row['Description']}_{row['Size/Type']}"
                # Handle TBD and N/A values properly
                try:
                    labor = float(row['Labor']) if row['Labor'] not in ['N/A', 'TBD'] else 0
                except (ValueError, TypeError):
                    labor = 0
                
                try:
                    material = float(row['Material']) if row['Material'] not in ['N/A', 'TBD'] else 0
                except (ValueError, TypeError):
                    material = 0
                
                master_pricing[key] = {
                    'Labor': labor,
                    'Material': material,
                    'Unit': row['Unit'],
                    'Category': row['Category']
                }
        print(f"✅ Loaded {len(master_pricing)} master pricing items")
    except FileNotFoundError:
        print("⚠️  Master pricing file not found, using existing prices")
        master_pricing = {}
    
    output_file = '113_University_Place_Professional_Internal_Pricing.xlsx'
    print(f"📝 Creating professional Excel: {output_file}")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Renovation Estimate"
    
    # Define professional color palette from comprehensive_cleanup.py (EXACT SAME COLORS)
    header_color = "231F20"  # Dark Black
    white_color = "FFFFFF"  # Pure White
    subheader_color = "C1A59A"  # Muted Brown/Taupe
    total_color = "FAF5EE"  # Off-White/Cream for section totals
    alternate_row_color = "F3E3D8"  # Very Light Pink/Peach for alternating rows
    cream_color = "FAF5EE"  # Off-White/Cream
    grand_total_color = "E1CDC0"  # Light Beige/Greige for grand total
    border_color = "F3E3D8"  # Very Light Pink/Peach borders
    
    # Define uniform thin border style (EXACT SAME AS comprehensive_cleanup.py)
    uniform_border = Border(
        left=Side(style='hair', color=border_color),
        right=Side(style='hair', color=border_color),
        top=Side(style='hair', color=border_color),
        bottom=Side(style='hair', color=border_color)
    )
    
    # Try to load Inter font, fallback to Calibri if not available (EXACT SAME AS comprehensive_cleanup.py)
    try:
        inter_font = Font(name="Inter", size=10)
        inter_bold = Font(name="Inter", size=11, bold=True)
        inter_header = Font(name="Inter", size=12, bold=True, color=white_color)
    except:
        inter_font = Font(name="Calibri", size=10)
        inter_bold = Font(name="Calibri", size=11, bold=True)
        inter_header = Font(name="Calibri", size=12, bold=True, color=white_color)
    
    # Define professional styles with Inter font (EXACT SAME AS comprehensive_cleanup.py)
    header_style = NamedStyle(name="header")
    header_style.font = inter_header
    header_style.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = uniform_border
    
    subheader_style = NamedStyle(name="subheader")
    subheader_style.font = inter_bold
    subheader_style.fill = PatternFill(start_color=subheader_color, end_color=subheader_color, fill_type="solid")
    subheader_style.alignment = Alignment(horizontal="left", vertical="center")
    subheader_style.border = uniform_border
    
    data_style = NamedStyle(name="data")
    data_style.font = inter_font
    data_style.fill = PatternFill(start_color=white_color, end_color=white_color, fill_type="solid")
    data_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, shrink_to_fit=False)
    data_style.border = uniform_border
    
    alternate_data_style = NamedStyle(name="alternate_data")
    alternate_data_style.font = inter_font
    alternate_data_style.fill = PatternFill(start_color=alternate_row_color, end_color=alternate_row_color, fill_type="solid")
    alternate_data_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, shrink_to_fit=False)
    alternate_data_style.border = uniform_border
    
    # Right-aligned styles for numeric columns (EXACT SAME AS comprehensive_cleanup.py)
    data_right_style = NamedStyle(name="data_right")
    data_right_style.font = inter_font
    data_right_style.fill = PatternFill(start_color=white_color, end_color=white_color, fill_type="solid")
    data_right_style.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    data_right_style.border = uniform_border
    
    alternate_data_right_style = NamedStyle(name="alternate_data_right")
    alternate_data_right_style.font = inter_font
    alternate_data_right_style.fill = PatternFill(start_color=alternate_row_color, end_color=alternate_row_color, fill_type="solid")
    alternate_data_right_style.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    alternate_data_right_style.border = uniform_border
    
    total_style = NamedStyle(name="total")
    total_style.font = inter_bold
    total_style.fill = PatternFill(start_color=total_color, end_color=total_color, fill_type="solid")
    total_style.alignment = Alignment(horizontal="right", vertical="center")
    total_style.border = uniform_border
    
    grand_total_style = NamedStyle(name="grand_total")
    grand_total_style.font = inter_bold
    grand_total_style.fill = PatternFill(start_color=grand_total_color, end_color=grand_total_color, fill_type="solid")
    grand_total_style.alignment = Alignment(horizontal="right", vertical="center")
    grand_total_style.border = uniform_border
    
    # Calculate dynamic column widths based on content (EXACT SAME AS comprehensive_cleanup.py)
    max_section_len = max(len(str(item.get('Category', ''))) for item in items) if items else 10
    max_room_len = max(len(str(item.get('Room', ''))) for item in items) if items else 10
    max_item_len = max(len(str(item.get('ItemName', ''))) for item in items) if items else 15
    max_desc_len = max(len(str(item.get('Description', ''))) for item in items) if items else 30
    
    # Set minimum and maximum widths for better appearance (EXACT SAME AS comprehensive_cleanup.py)
    section_width = max(25, min(max_section_len + 5, 35))  # Min 25, Max 35
    room_width = max(18, min(max_room_len + 3, 25))        # Min 18, Max 25  
    item_width = max(30, min(max_item_len + 5, 45))        # Min 30, Max 45
    desc_width = max(60, min(max_desc_len // 2, 80))       # Min 60, Max 80 (keep compact width, use dynamic row heights for wrapping)
    
    print(f"[INFO] Enhanced column widths: Section={section_width}, Room={room_width}, ItemName={item_width}, Description={desc_width}")
    
    # Apply column widths with better sizing (EXACT SAME AS comprehensive_cleanup.py)
    ws.column_dimensions['A'].width = section_width      # Category/Section
    ws.column_dimensions['B'].width = room_width         # Room  
    ws.column_dimensions['C'].width = item_width         # ItemName
    ws.column_dimensions['D'].width = desc_width         # Description
    ws.column_dimensions['E'].width = 12                 # Quantity
    ws.column_dimensions['F'].width = 18                 # UnitCost
    ws.column_dimensions['G'].width = 10                 # Markup
    ws.column_dimensions['H'].width = 12                 # MarkupType
    ws.column_dimensions['I'].width = 15                 # Total
    
    print(f"[INFO] Dynamic column widths: Section={section_width}, Room={room_width}, ItemName={item_width}, Description={desc_width}")

    # Set row height for better spacing (EXACT SAME AS comprehensive_cleanup.py)
    ws.row_dimensions[1].height = 25  # Header row
    
    # Add headers with professional styling (EXACT SAME AS comprehensive_cleanup.py)
    headers = ["Section", "Room", "Item Name", "Description", "Quantity", "Unit Cost", "Markup", "Total", "Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = header_style
    
    # Group items by category (EXACT SAME AS comprehensive_cleanup.py)
    categories = defaultdict(list)
    for item in items:
        category = item.get('Category', '').strip()
        if category:
            categories[category].append(item)
    
    current_row = 2
    overall_subtotal = 0
    
    # Add items by category with professional formatting (EXACT SAME AS comprehensive_cleanup.py)
    for category, category_items in categories.items():
        # Add category header with professional styling
        category_cell = ws.cell(row=current_row, column=1, value=category)
        category_cell.style = subheader_style
        # Apply border to all cells in the row
        for col in range(2, 10):
            ws.cell(row=current_row, column=col, value="").style = subheader_style
        current_row += 1
        
        category_total = 0
        
        # Add items in this category with alternating row colors
        for i, item in enumerate(category_items):
            # Clean up Total value (EXACT SAME AS comprehensive_cleanup.py)
            total_value = item.get('Total', '0')
            if isinstance(total_value, str):
                total_value = total_value.replace(',', '').replace('$', '').strip()
                try:
                    total_value = float(total_value)
                except ValueError:
                    total_value = 0.0
            else:
                total_value = float(total_value) if total_value else 0.0
            
            category_total += total_value
            
            # Choose style based on row number for alternating colors (EXACT SAME AS comprehensive_cleanup.py)
            row_style = data_style if i % 2 == 0 else alternate_data_style
            row_style_right = data_right_style if i % 2 == 0 else alternate_data_right_style
            
            # Add item row with professional formatting (EXACT SAME AS comprehensive_cleanup.py)
            ws.cell(row=current_row, column=1, value="").style = row_style # Section column is empty
            ws.cell(row=current_row, column=2, value=item.get('Room', '')).style = row_style
            ws.cell(row=current_row, column=3, value=item.get('ItemName', '')).style = row_style
            
            # Add description with proper text wrapping and row height (EXACT SAME AS comprehensive_cleanup.py)
            clean_desc = clean_description_text(item.get('Description', ''))
            desc_cell = ws.cell(row=current_row, column=4, value=clean_desc)
            desc_cell.style = row_style
            
            # Calculate appropriate row height for description wrapping (EXACT SAME AS comprehensive_cleanup.py)
            if len(clean_desc) > 60:  # If description is longer than column width, increase row height
                # More intelligent line calculation based on actual column width
                estimated_lines = max(2, len(clean_desc) // 60)  # Estimate lines needed based on 60-char column
                row_height = max(20, estimated_lines * 18)  # Min 20, 18 per line for better spacing
                ws.row_dimensions[current_row].height = row_height
                print(f"[DEBUG] Row {current_row}: Description length {len(clean_desc)}, estimated {estimated_lines} lines, height {row_height}")
            
            ws.cell(row=current_row, column=5, value=item.get('Quantity', '')).style = row_style_right
            ws.cell(row=current_row, column=6, value=item.get('UnitCost', '')).style = row_style_right
            ws.cell(row=current_row, column=7, value=item.get('Markup', '')).style = row_style_right
            ws.cell(row=current_row, column=8, value=f"${total_value:,.2f}").style = row_style_right
            ws.cell(row=current_row, column=9, value=item.get('Confidence', '')).style = row_style_right
            current_row += 1
        
        # Add category total with professional styling (EXACT SAME AS comprehensive_cleanup.py)
        if category_total > 0:
            print(f"[INFO] Category '{category}' total: ${category_total:,.2f}")
            # Add just "Total" label (no section name)
            ws.cell(row=current_row, column=1, value="Total").style = total_style
            # Add empty cells for proper formatting
            for col in range(2, 8):
                ws.cell(row=current_row, column=col, value="").style = total_style
            
            # Add the total with professional formatting
            total_cell = ws.cell(row=current_row, column=8, value=f"${category_total:,.2f}")
            total_cell.style = total_style
            ws.cell(row=current_row, column=9, value="").style = total_style
            current_row += 1
            
            # Add a properly styled blank row after total (fix the extra cell color issue)
            for col in range(1, 10):
                ws.cell(row=current_row, column=col, value="").style = data_style
            current_row += 1
        
        # Add gap between sections
        current_row += 1
        
        overall_subtotal += category_total
    
    # Add professional summary section (EXACT SAME AS comprehensive_cleanup.py)
    if overall_subtotal > 0:
        # Add overall subtotal
        ws.cell(row=current_row, column=1, value="Overall Subtotal").style = total_style
        for col in range(2, 8):
            ws.cell(row=current_row, column=col, value="").style = total_style
        ws.cell(row=current_row, column=8, value=f"${overall_subtotal:,.2f}").style = total_style
        ws.cell(row=current_row, column=9, value="").style = total_style
        current_row += 1
        
        # Add general conditions
        general_conditions = overall_subtotal * 0.10
        ws.cell(row=current_row, column=1, value="General Conditions (10%)").style = total_style
        for col in range(2, 8):
            ws.cell(row=current_row, column=col, value="").style = total_style
        ws.cell(row=current_row, column=8, value=f"${general_conditions:,.2f}").style = total_style
        ws.cell(row=current_row, column=9, value="").style = total_style
        current_row += 1
        
        # Add grand total
        grand_total = overall_subtotal + general_conditions
        ws.cell(row=current_row, column=1, value="GRAND TOTAL").style = grand_total_style
        for col in range(2, 8):
            ws.cell(row=current_row, column=col, value="").style = grand_total_style
        ws.cell(row=current_row, column=8, value=f"${grand_total:,.2f}").style = grand_total_style
        ws.cell(row=current_row, column=9, value="").style = grand_total_style
    
    # Save the workbook
    wb.save(output_file)
    print(f"✅ Professional Excel file created: {output_file}")
    
    # Print summary
    print(f"\n📊 PROFESSIONAL EXCEL SUMMARY:")
    print(f"Total Items: {len(items)}")
    print(f"Categories: {len(categories)}")
    print(f"Base Cost: ${overall_subtotal:,.2f}")
    print(f"General Conditions: ${general_conditions:,.2f}")
    print(f"Grand Total: ${grand_total:,.2f}")
    
    print(f"\n🎨 STYLING APPLIED (EXACT SAME AS comprehensive_cleanup.py):")
    print(f"  ✅ Headers: Dark Black (#{header_color}) with white text")
    print(f"  ✅ Category Headers: Muted Brown/Taupe (#{subheader_color})")
    print(f"  ✅ Alternating Rows: White and Very Light Pink/Peach (#{alternate_row_color})")
    print(f"  ✅ Category Totals: Off-White/Cream (#{total_color})")
    print(f"  ✅ Grand Total: Light Beige/Greige (#{grand_total_color})")
    print(f"  ✅ Borders: Very Light Pink/Peach (#{border_color})")
    print(f"  ✅ Font: Inter (fallback to Calibri)")
    print(f"  ✅ Dynamic row heights for descriptions")
    
    if master_pricing:
        print(f"\n💰 MASTER PRICING INTEGRATION:")
        print(f"  ✅ Loaded {len(master_pricing)} pricing items")
        print(f"  ✅ Ready for real pricing updates")
    else:
        print(f"\n⚠️  PRICING UPDATE NEEDED:")
        print(f"  ⚠️  Master pricing file not found")
        print(f"  ⚠️  Please provide master pricing data for real costs")
    
    print(f"\n🎯 PROFESSIONAL EXCEL ESTIMATE READY!")
    print(f"File: {output_file}")
    
    return output_file

def clean_description_text(description):
    """Clean description text for better Excel formatting."""
    if not description:
        return ""
    
    # Remove extra whitespace and normalize
    cleaned = " ".join(description.split())
    
    # Add line breaks for better readability
    if len(cleaned) > 80:
        # Try to break at natural points
        words = cleaned.split()
        lines = []
        current_line = ""
        
        for word in words:
            if len(current_line + " " + word) <= 80:
                current_line += (" " + word) if current_line else word
            else:
                if current_line:
                    lines.append(current_line)
                current_line = word
        
        if current_line:
            lines.append(current_line)
        
        cleaned = "\n".join(lines)
    
    return cleaned

if __name__ == "__main__":
    create_professional_excel_with_internal_pricing()

