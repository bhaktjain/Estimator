#!/usr/bin/env python3
"""Convert 366 Broadway estimate CSV to Excel."""
import csv
import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from collections import defaultdict

def read_items(csv_path):
    items = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get('ItemName') and row.get('Category') and row['Category'] != 'Category':
                items.append(row)
    return items

def clean_total(val):
    if isinstance(val, str):
        val = val.replace(',', '').replace('$', '').strip()
    try:
        return float(val) if val else 0.0
    except ValueError:
        return 0.0

def create_excel(items, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "366 Broadway Estimate"
    
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    total_font = Font(bold=True, size=11)
    
    headers = ["Section", "Room", "Item Name", "Description", "Quantity", "Unit Cost", "Markup", "Total", "Confidence"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
    ws.row_dimensions[1].height = 22
    
    categories = defaultdict(list)
    for item in items:
        cat = item.get('Category', '').strip()
        if cat:
            categories[cat].append(item)
    
    order = ['Demolition', 'Walls & Ceiling', 'Plumbing', 'Waterproofing', 'Tile', 'Bathroom Fixtures',
             'Electrical', 'Flooring', 'Painting & Wall Coverings', 'Trims', 'Cabinetry & Storage',
             'Countertops', 'Backsplash', 'Appliances', 'Stairs', 'Radiator Covers', 'Doors', 'General Requirements']
    ordered = [c for c in order if c in categories]
    for c in sorted(categories):
        if c not in ordered:
            ordered.append(c)
    
    row = 2
    subtotal = 0
    for cat in ordered:
        for item in categories[cat]:
            total_val = clean_total(item.get('Total', 0))
            subtotal += total_val
            ws.cell(row=row, column=1, value="")
            ws.cell(row=row, column=2, value=item.get('Room', ''))
            ws.cell(row=row, column=3, value=item.get('ItemName', ''))
            ws.cell(row=row, column=4, value=item.get('Description', ''))
            ws.cell(row=row, column=5, value=item.get('Quantity', ''))
            ws.cell(row=row, column=6, value=item.get('UnitCost', ''))
            ws.cell(row=row, column=7, value=item.get('Markup', ''))
            ws.cell(row=row, column=8, value=f"${total_val:,.2f}")
            ws.cell(row=row, column=9, value=item.get('Confidence', ''))
            row += 1
    
    cat_total = sum(clean_total(i.get('Total', 0)) for i in items)
    gc = cat_total * 0.10
    grand = cat_total + gc
    
    ws.cell(row=row, column=1, value="Subtotal").fill = total_fill
    ws.cell(row=row, column=8, value=f"${cat_total:,.2f}").fill = total_fill
    row += 1
    ws.cell(row=row, column=1, value="General Conditions (10%)").fill = total_fill
    ws.cell(row=row, column=8, value=f"${gc:,.2f}").fill = total_fill
    row += 1
    ws.cell(row=row, column=1, value="Grand Total").fill = total_fill
    ws.cell(row=row, column=8, value=f"${grand:,.2f}").fill = total_fill
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 12
    
    wb.save(output_path)
    print(f"Excel saved: {output_path}")
    print(f"Subtotal: ${cat_total:,.2f}")
    print(f"Grand Total: ${grand:,.2f}")

if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    csv_path = os.path.join(base, 'estimate_366_broadway.csv')
    excel_path = os.path.join(base, '366_Broadway_Gambhir_Estimate_Final.xlsx')
    items = read_items(csv_path)
    print(f"Read {len(items)} items")
    create_excel(items, excel_path)
