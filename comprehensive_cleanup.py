import csv
import json
from collections import defaultdict
import re
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows

def read_csv_items(file_path):
    """Read CSV and return items as list of dictionaries."""
    items = []
    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row['Category'] and row['ItemName'] and row['Category'] != 'Category':
                items.append(row)
    return items

def get_valid_sections():
    """Get valid sections from section_minimums_margins.csv."""
    valid_sections = set()
    with open('section_minimums_margins.csv', 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            valid_sections.add(row['Section'].strip())
    return valid_sections

def validate_pricing_data():
    """Validate that items use actual pricing codes from master pricing sheet."""
    valid_pricing_codes = set()
    try:
        with open('master_pricing_data.csv', 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                item_code = row.get('Item Code', '').strip()
                if item_code:
                    valid_pricing_codes.add(item_code)
    except:
        print("[WARNING] Could not read master pricing data")
        return set()
    return valid_pricing_codes

def get_pricing_sheet_sections():
    """Get all sections from the master pricing sheet."""
    pricing_sections = set()
    try:
        with open('master_pricing_data.csv', 'r', encoding='utf-8') as f:
            import csv
            reader = csv.DictReader(f)
            for row in reader:
                category = row.get('Category', '').strip()
                if category:
                    pricing_sections.add(category)
    except:
        # If we can't read the pricing sheet, use a default set
        pricing_sections = {
            'Demolition', 'Electrical', 'Plumbing', 'Kitchen Cabinets', 'Kitchen',
            'Tile', 'Carpentry', 'Waterproofing', 'Cleaning'
        }
    return pricing_sections

def normalize_item_name(item_name):
    """Normalize item name for comparison."""
    return re.sub(r'[^\w\s]', '', item_name.lower().strip())

def normalize_room_name(room_name):
    """Normalize room name for comparison."""
    room = room_name.lower().strip()
    # Standardize room names but preserve bathroom distinctions
    if 'primary bathroom' in room or 'main bathroom' in room:
        return 'primary bathroom'
    elif 'secondary bathroom' in room or 'second bathroom' in room or 'guest bathroom' in room:
        return 'secondary bathroom'
    elif 'bathroom 2' in room or 'bathroom2' in room:
        return 'bathroom 2'
    elif 'bathroom 1' in room or 'bathroom1' in room:
        return 'bathroom 1'
    elif 'bathroom' in room:
        return 'bathroom'
    elif 'kitchen' in room or 'other 1' in room or 'other1' in room:
        return 'kitchen'
    elif 'closet' in room:
        return 'closet'
    elif 'living' in room or 'room' in room:
        return 'living area'
    elif 'office' in room:
        return 'office area'
    elif 'entry' in room or 'foyer' in room:
        return 'entry'
    elif 'general' in room or 'apartment' in room or 'full apartment' in room:
        return 'apartment'
    else:
        return room

def is_same_work(item1, item2):
    """Check if two items represent the same work."""
    name1 = normalize_item_name(item1.get('ItemName', ''))
    name2 = normalize_item_name(item2.get('ItemName', ''))
    room1 = normalize_room_name(item1.get('Room', ''))
    room2 = normalize_room_name(item2.get('Room', ''))
    
    # If different rooms, not the same work
    if room1 != room2:
        return False
    
    # Check for exact name match
    if name1 == name2:
        return True
    
    # Check for high similarity
    words1 = set(name1.split())
    words2 = set(name2.split())
    
    if len(words1) > 0 and len(words2) > 0:
        similarity = len(words1.intersection(words2)) / max(len(words1), len(words2))
        return similarity > 0.7
    
    return False

def is_overlapping_work(item1, item2):
    """Check if two items represent overlapping work."""
    name1 = normalize_item_name(item1.get('ItemName', ''))
    name2 = normalize_item_name(item2.get('ItemName', ''))
    room1 = normalize_room_name(item1.get('Room', ''))
    room2 = normalize_room_name(item2.get('Room', ''))
    
    # If different rooms, not overlapping
    if room1 != room2:
        return False
    
    # Check for overlapping keywords
    painting_keywords = ['paint', 'painting', 'primer', 'coat', 'finish']
    wall_keywords = ['wall', 'drywall', 'sheetrock', 'plaster']
    floor_keywords = ['floor', 'flooring', 'hardwood', 'tile', 'vinyl']
    electrical_keywords = ['electrical', 'wiring', 'outlet', 'switch', 'light']
    plumbing_keywords = ['plumbing', 'pipe', 'sink', 'toilet', 'shower']
    
    # Check if both items contain overlapping keywords
    name1_has_painting = any(keyword in name1 for keyword in painting_keywords)
    name2_has_painting = any(keyword in name2 for keyword in painting_keywords)
    
    name1_has_wall = any(keyword in name1 for keyword in wall_keywords)
    name2_has_wall = any(keyword in name2 for keyword in wall_keywords)
    
    name1_has_floor = any(keyword in name1 for keyword in floor_keywords)
    name2_has_floor = any(keyword in name2 for keyword in floor_keywords)
    
    name1_has_electrical = any(keyword in name1 for keyword in electrical_keywords)
    name2_has_electrical = any(keyword in name2 for keyword in electrical_keywords)
    
    name1_has_plumbing = any(keyword in name1 for keyword in plumbing_keywords)
    name2_has_plumbing = any(keyword in name2 for keyword in plumbing_keywords)
    
    # If both have the same type of work, they overlap
    if (name1_has_painting and name2_has_painting) or \
       (name1_has_wall and name2_has_wall) or \
       (name1_has_floor and name2_has_floor) or \
       (name1_has_electrical and name2_has_electrical) or \
       (name1_has_plumbing and name2_has_plumbing):
        return True
    
    return False

def recategorize_items(items):
    """Recategorize items based on reference CSV patterns and trade-based categorization."""
    print(f"[INFO] Recategorizing {len(items)} items using reference CSV patterns")
    
    recategorized = []
    
    for item in items:
        name = item.get('ItemName', '').lower()
        description = item.get('Description', '').lower()
        room = item.get('Room', '').lower()
        category = item.get('Category', '').strip()
        old_category = category
        
        # Based on reference CSV analysis, categorize by exact patterns found
        # Order matters - DEMOLITION MUST BE FIRST to catch "gut", "remove", etc.
        new_category = None
        
        # HIGHEST PRIORITY - DEMOLITION (must come first!)
        if ('gut' in name or 'gut' in description or
            'demolition' in name or 'demolition' in description or
            'demo' in name or 'demo' in description or
            'remove all' in name or 'remove all' in description or
            'remove existing' in name or 'remove existing' in description or
            'tear out' in name or 'tear out' in description or
            'strip' in name or 'strip' in description or
            'pre-construction' in name or 'pre-construction' in description) and 'flooring' not in name and 'flooring' not in description:
            new_category = 'Demolition'
        
        # SECOND PRIORITY - Very specific items (only if not demolition)
        elif 'countertop' in name or ('countertop' in description and 'remove' not in description):
            new_category = 'Countertops'
        elif 'backsplash' in name or ('backsplash' in description and 'remove' not in description):
            new_category = 'Backsplash'
        elif ('appliance' in name or 'appliance' in description or
              'dishwasher' in name or 'refrigerator' in name or 'stove' in name) and 'remove' not in description and 'electrical' not in name:
            new_category = 'Appliances'
        elif ('cabinet' in name or 'cabinet' in description) and 'remove' not in description and 'gut' not in description:
            new_category = 'Cabinetry'
        
        # THIRD PRIORITY - Trade-specific work
        elif ('plumbing' in name or 'plumbing' in description or
              'sink' in name or 'toilet' in name or 'shower' in name or 'tub' in name or
              'fixture' in name or 'faucet' in name or 'drain' in name or
              'water line' in name or 'waste line' in name):
            new_category = 'Plumbing'
        elif ('electrical' in name or 'electrical' in description or
              'wiring' in name or 'outlet' in name or 'switch' in name or
              'light' in name or 'gfi' in name or 'circuit' in name):
            new_category = 'Electrical'
        elif ('waterproof' in name or 'waterproof' in description or
              'moisture' in name or 'membrane' in name or 'vapor barrier' in name):
            new_category = 'Waterproofing'
        elif ('tile' in name or 'tile' in description and 'backsplash' not in name):
            new_category = 'Tile'
        
        # FOURTH PRIORITY - Construction work
        elif ('drywall' in name or 'drywall' in description or
              'wall' in name and 'tile' not in name or 'ceiling' in name or
              'soffit' in name or 'partition' in name or 'framing' in name):
            new_category = 'Walls & Ceiling'
        elif ('flooring' in name or 'flooring' in description or
              'hardwood' in name or 'laminate' in name or 'vinyl' in name or
              ('floor' in name and 'tile' not in name and 'demo' not in name)):
            new_category = 'Flooring'
        elif ('paint' in name or 'painting' in name or 'primer' in name):
            new_category = 'Painting & Wall Coverings'
        elif ('trim' in name or 'baseboard' in name or 'molding' in name or 'crown' in name):
            new_category = 'Trims'
        elif ('door' in name or 'pocket door' in name):
            new_category = 'Doors'
        elif ('window' in name or 'glazing' in name):
            new_category = 'Windows'
        
        # FIFTH PRIORITY - Systems
        elif ('hvac' in name or 'heating' in name or 'cooling' in name or 'radiator' in name):
            new_category = 'Heating and Cooling'
        elif ('accessory' in name or 'towel bar' in name or 'mirror' in name or
              'medicine cabinet' in name):
            new_category = 'Accessories'
        elif ('cleaning' in name or 'cleanup' in name or 'final' in name or
              'general conditions' in name or 'conditions' in name):
            new_category = 'General Requirements'
        
        # Apply the new category if determined
        if new_category and new_category != old_category:
            item['Category'] = new_category
            print(f"[INFO] Recategorized '{name}' from {old_category} to {new_category}")
        
        recategorized.append(item)
    
    return recategorized

def is_countertop_duplicate(item1, item2):
    """Check if two countertop items are duplicates."""
    name1 = normalize_item_name(item1.get('ItemName', ''))
    name2 = normalize_item_name(item2.get('ItemName', ''))
    room1 = normalize_room_name(item1.get('Room', ''))
    room2 = normalize_room_name(item2.get('Room', ''))
    
    # If rooms are different, they're not duplicates (allow different rooms)
    if room1 != room2:
        return False
    
    # Check for similar countertop work
    countertop_keywords = ['countertop', 'counter', 'quartz', 'marble', 'granite', 'surface']
    if any(keyword in name1.lower() for keyword in countertop_keywords) and any(keyword in name2.lower() for keyword in countertop_keywords):
        return True
    
    return False

def is_electrical_duplicate(item1, item2):
    """Check if two electrical items are duplicates."""
    name1 = normalize_item_name(item1.get('ItemName', ''))
    name2 = normalize_item_name(item2.get('ItemName', ''))
    room1 = normalize_room_name(item1.get('Room', ''))
    room2 = normalize_room_name(item2.get('Room', ''))
    
    # If rooms are different, they're not duplicates (allow different rooms)
    if room1 != room2:
        return False
    
    # Check for similar electrical work
    electrical_keywords = ['electrical', 'wiring', 'outlet', 'switch', 'light', 'panel', 'rewiring']
    if any(keyword in name1.lower() for keyword in electrical_keywords) and any(keyword in name2.lower() for keyword in electrical_keywords):
        return True
    
    return False

def is_plumbing_duplicate(item1, item2):
    """Check if two plumbing items are duplicates."""
    room1 = normalize_room_name(item1.get('Room', ''))
    room2 = normalize_room_name(item2.get('Room', ''))
    item1_name = normalize_item_name(item1.get('ItemName', ''))
    item2_name = normalize_item_name(item2.get('ItemName', ''))
    
    # If rooms are different, they're not duplicates (allow different rooms)
    if room1 != room2:
        return False
    
    # Check for overlapping scope
    if 'full gut' in item1_name.lower() and 'full gut' in item2_name.lower():
        return True
    
    # Check for bathroom vs individual bathroom conflicts
    if (room1 == 'bathrooms' and room2 in ['bathroom 1', 'bathroom 2']) or \
       (room2 == 'bathrooms' and room1 in ['bathroom 1', 'bathroom 2']):
        # If one is "bathrooms" (both) and other is individual, keep the "bathrooms" one
        if 'bathrooms' in [room1, room2]:
            return True
    
    # Check for similar item names
    if is_same_work(item1, item2):
        return True
    
    return False

def is_demolition_duplicate(item1, item2):
    """Check if two demolition items are duplicates."""
    name1 = normalize_item_name(item1.get('ItemName', ''))
    name2 = normalize_item_name(item2.get('ItemName', ''))
    room1 = normalize_room_name(item1.get('Room', ''))
    room2 = normalize_room_name(item2.get('Room', ''))
    
    # If rooms are different, they're not duplicates (allow different rooms)
    if room1 != room2:
        return False
    
    # Check for similar demolition work
    demo_keywords = ['demolition', 'demo', 'gut', 'remove', 'tear out', 'strip']
    if any(keyword in name1.lower() for keyword in demo_keywords) and any(keyword in name2.lower() for keyword in demo_keywords):
        return True
    
    return False

def is_painting_duplicate(item1, item2):
    """Check if two painting items are duplicates."""
    name1 = normalize_item_name(item1.get('ItemName', ''))
    name2 = normalize_item_name(item2.get('ItemName', ''))
    room1 = normalize_room_name(item1.get('Room', ''))
    room2 = normalize_room_name(item2.get('Room', ''))
    
    # If rooms are different, they're not duplicates (allow different rooms)
    if room1 != room2:
        return False
    
    # Check for similar painting work
    painting_keywords = ['paint', 'painting', 'primer', 'coat', 'finish']
    if any(keyword in name1.lower() for keyword in painting_keywords) and any(keyword in name2.lower() for keyword in painting_keywords):
        return True
    
    return False

def is_tile_duplicate(item1, item2):
    """Check if two tile items are duplicates."""
    name1 = normalize_item_name(item1.get('ItemName', ''))
    name2 = normalize_item_name(item2.get('ItemName', ''))
    room1 = normalize_room_name(item1.get('Room', ''))
    room2 = normalize_room_name(item2.get('Room', ''))
    
    # If rooms are different, they're not duplicates (allow different rooms)
    if room1 != room2:
        return False
    
    # Check for bathroom vs individual bathroom conflicts
    if (room1 == 'bathrooms' and room2 in ['bathroom 1', 'bathroom 2']) or \
       (room2 == 'bathrooms' and room1 in ['bathroom 1', 'bathroom 2']):
        # If one is "bathrooms" (both) and other is individual, keep the "bathrooms" one
        if 'bathrooms' in [room1, room2]:
            return True
    
    # Check for similar tile work
    tile_keywords = ['tile', 'backsplash', 'grout', 'ceramic', 'porcelain']
    desc1 = normalize_item_name(item1.get('Description', ''))
    desc2 = normalize_item_name(item2.get('Description', ''))
    
    # Check both name and description for tile keywords
    combined1 = f"{name1} {desc1}".lower()
    combined2 = f"{name2} {desc2}".lower()
    
    if any(keyword in combined1 for keyword in tile_keywords) and any(keyword in combined2 for keyword in tile_keywords):
        return True
    
    # Special case: kitchen flooring conflicts between tile and flooring categories
    if 'kitchen' in room1.lower() and 'kitchen' in room2.lower():
        if any(keyword in combined1 for keyword in ['floor', 'flooring']) and any(keyword in combined2 for keyword in ['tile', 'floor', 'flooring']):
            return True
    
    return False

def enhanced_deduplication(items):
    """Enhanced deduplication to ensure truly unique, non-overlapping items per section."""
    print(f"[INFO] Performing enhanced deduplication on {len(items)} items")
    
    # Group items by category
    categories = defaultdict(list)
    for item in items:
        category = item.get('Category', '').strip()
        if category:
            categories[category].append(item)
    
    cleaned_items = []
    
    for category, category_items in categories.items():
        print(f"[INFO] Processing category: {category} with {len(category_items)} items")
        
        # Sort by confidence score (highest first) and then by room specificity
        def sort_key(item):
            confidence = float(item.get('Confidence', 0))
            room = item.get('Room', '').lower()
            # Prioritize specific rooms over general ones
            specificity = 0
            if 'bathroom 1' in room or 'bathroom 2' in room or 'kitchen' in room:
                specificity = 2
            elif 'bathrooms' in room or 'entire apartment' in room:
                specificity = 1
            return (-confidence, -specificity)
        
        category_items.sort(key=sort_key)
        
        # Track unique items in this category
        unique_items = []
        seen_combinations = set()
        
        for item in category_items:
            # Create a unique identifier for this item
            room = item.get('Room', '').strip().lower()
            item_name = item.get('ItemName', '').strip().lower()
            description = item.get('Description', '').strip().lower()
            
            # Create different combinations for duplicate detection
            combinations = [
                f"{room}|{item_name}",
                f"{room}|{item_name}|{description}",
                f"{item_name}|{description}"
            ]
            
            # Check if this item is a duplicate
            is_duplicate = False
            for combo in combinations:
                if combo in seen_combinations:
                    print(f"[INFO] Removing duplicate: {item.get('ItemName', '')} in {item.get('Room', '')}")
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                # Add all combinations to seen set
                for combo in combinations:
                    seen_combinations.add(combo)
                
                # Apply category-specific duplicate checks
                should_include = True
                
                # Check for specific duplicate types (only within same room)
                for existing_item in unique_items:
                    # Only check for duplicates if they're in the same room
                    if item.get('Room', '').strip().lower() == existing_item.get('Room', '').strip().lower():
                        if is_countertop_duplicate(item, existing_item):
                            print(f"[INFO] Removing countertop duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_plumbing_duplicate(item, existing_item):
                            print(f"[INFO] Removing plumbing duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_demolition_duplicate(item, existing_item):
                            print(f"[INFO] Removing demolition duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_painting_duplicate(item, existing_item):
                            print(f"[INFO] Removing painting duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_tile_duplicate(item, existing_item):
                            print(f"[INFO] Removing tile duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_commercial_cleaning_duplicate(item, existing_item):
                            print(f"[INFO] Removing commercial cleaning duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_flooring_duplicate(item, existing_item):
                            print(f"[INFO] Removing flooring duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_waterproofing_duplicate(item, existing_item):
                            print(f"[INFO] Removing waterproofing duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_walls_ceilings_duplicate(item, existing_item):
                            print(f"[INFO] Removing walls/ceilings duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_backsplash_duplicate(item, existing_item):
                            print(f"[INFO] Removing backsplash duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_trim_duplicate(item, existing_item):
                            print(f"[INFO] Removing trim duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_doors_duplicate(item, existing_item):
                            print(f"[INFO] Removing doors duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                        elif is_cabinetry_duplicate(item, existing_item):
                            print(f"[INFO] Removing cabinetry duplicate in same room: {item.get('ItemName', '')} in {item.get('Room', '')}")
                            should_include = False
                            break
                
                if should_include:
                    unique_items.append(item)
                    print(f"[INFO] Keeping unique item: {item.get('ItemName', '')} in {item.get('Room', '')}")
        
        print(f"[INFO] Category '{category}': {len(category_items)} -> {len(unique_items)} items")
        cleaned_items.extend(unique_items)
    
    print(f"[INFO] Enhanced deduplication complete: {len(items)} -> {len(cleaned_items)} items")
    return cleaned_items

def prioritize_high_confidence(items):
    """Prioritize items with higher confidence scores when there's ambiguity."""
    print(f"[INFO] Prioritizing high-confidence items from {len(items)} items")
    
    # Group items by category and room
    grouped_items = {}
    for item in items:
        category = item.get('Category', '').strip()
        room = normalize_room_name(item.get('Room', ''))
        item_name = normalize_item_name(item.get('ItemName', ''))
        
        key = f"{category}_{room}_{item_name}"
        if key not in grouped_items:
            grouped_items[key] = []
        grouped_items[key].append(item)
    
    prioritized_items = []
    
    for key, similar_items in grouped_items.items():
        if len(similar_items) == 1:
            # No ambiguity, keep the item
            prioritized_items.append(similar_items[0])
        else:
            # Multiple similar items, prioritize by confidence
            # Convert confidence to float for comparison
            for item in similar_items:
                try:
                    confidence = float(item.get('Confidence', '0').replace('%', ''))
                    item['_confidence_float'] = confidence
                except:
                    item['_confidence_float'] = 0
            
            # Sort by confidence (highest first)
            similar_items.sort(key=lambda x: x['_confidence_float'], reverse=True)
            
            # Keep the highest confidence item
            best_item = similar_items[0]
            confidence_value = best_item['_confidence_float']  # Store before deleting
            del best_item['_confidence_float']  # Clean up temporary field
            
            if confidence_value >= 85:
                prioritized_items.append(best_item)
                print(f"[INFO] Selected high-confidence item: {best_item.get('ItemName', '')} (confidence: {confidence_value})")
            else:
                print(f"[INFO] Skipped low-confidence item: {best_item.get('ItemName', '')} (confidence: {confidence_value})")
    
    print(f"[INFO] After prioritizing: {len(prioritized_items)} items")
    return prioritized_items

def comprehensive_cleanup(items):
    """Perform comprehensive cleanup on estimate items."""
    print(f"[INFO] Starting comprehensive cleanup with {len(items)} items")
    
    # Basic cleaning
    cleaned_items = []
    for item in items:
        # Clean up item names and descriptions
        item['ItemName'] = item.get('ItemName', '').strip()
        item['Description'] = clean_description_text(item.get('Description', ''))
        item['Room'] = item.get('Room', '').strip()
        item['Category'] = item.get('Category', '').strip()
        
        # Only include items with valid names
        if item['ItemName']:
            cleaned_items.append(item)
    
    print(f"[INFO] After basic cleaning: {len(cleaned_items)} items")
    
    # Recategorize items
    cleaned_items = recategorize_items(cleaned_items)
    print(f"[INFO] After recategorization: {len(cleaned_items)} items")
    
    # GPT-based deduplication (if API key is available)
    try:
        import os
        api_key = os.getenv('OPENAI_API_KEY')
        if api_key:
            from gpt_deduplication import gpt_deduplication
            cleaned_items = gpt_deduplication(cleaned_items, api_key)
            print(f"[INFO] After GPT deduplication: {len(cleaned_items)} items")
        else:
            print("[INFO] No OpenAI API key found, using local deduplication")
            # Fall back to local deduplication
            cleaned_items = enhanced_deduplication(cleaned_items)
            print(f"[INFO] After local deduplication: {len(cleaned_items)} items")
    except Exception as e:
        print(f"[WARNING] GPT deduplication failed: {e}, using local deduplication")
        cleaned_items = enhanced_deduplication(cleaned_items)
        print(f"[INFO] After local deduplication: {len(cleaned_items)} items")
    
    # Prioritize high-confidence items
    cleaned_items = prioritize_high_confidence(cleaned_items)
    print(f"[INFO] After prioritizing confidence: {len(cleaned_items)} items")
    
    # Fix total values (formula evaluation)
    cleaned_items = fix_total_values(cleaned_items)
    print(f"[INFO] After fixing total values: {len(cleaned_items)} items")
    
    # Remove smaller room items when full apartment scope exists for the same category
    cleaned_items = remove_smaller_rooms_when_full_apartment_exists(cleaned_items)
    print(f"[INFO] After removing smaller rooms: {len(cleaned_items)} items")
    
    # Merge overlapping items to eliminate duplicates
    cleaned_items = merge_overlapping_items(cleaned_items)
    print(f"[INFO] After merging overlapping items: {len(cleaned_items)} items")
    
    # Remove cross-category duplicates (same work in different categories)
    cleaned_items = remove_cross_category_duplicates(cleaned_items)
    print(f"[INFO] After removing cross-category duplicates: {len(cleaned_items)} items")
    
    # Merge cabinetry categories
    cleaned_items = merge_cabinetry_categories(cleaned_items)
    print(f"[INFO] After merging cabinetry categories: {len(cleaned_items)} items")
    
    print(f"[INFO] After comprehensive cleanup: {len(cleaned_items)} items")
    return cleaned_items

def clean_description_text(description):
    """Clean up description text to remove awkward line breaks and improve readability."""
    if not description:
        return ""
    
    # Convert to string if needed
    desc = str(description)
    
    # Remove multiple spaces
    desc = ' '.join(desc.split())
    
    # Remove awkward line breaks that might have been added during processing
    desc = desc.replace('\n', ' ').replace('\r', ' ')
    
    # Remove multiple spaces again
    desc = ' '.join(desc.split())
    
    # Try to break at better boundaries for Excel wrapping
    # If description is very long, try to break at sentence boundaries
    if len(desc) > 120:
        # Look for natural break points like commas, periods, or "and"
        break_points = [',', '.', ' and ', ' or ', ' with ', ' including ']
        for break_point in break_points:
            if break_point in desc:
                # Don't break if it would make lines too short
                parts = desc.split(break_point)
                if all(len(part.strip()) > 20 for part in parts):
                    # This is a good break point, keep as is
                    break
                else:
                    # This break would create very short lines, don't use it
                    continue
    
    return desc

def create_excel_file(items, output_file):
    """Create a beautifully formatted Excel file with the cleaned estimate data."""
    print(f"[INFO] Creating Excel file: {output_file}")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Renovation Estimate"
    
    # Define professional color palette from the reference image
    header_color = "231F20"  # Dark Black
    white_color = "FFFFFF"  # Pure White
    subheader_color = "C1A59A"  # Muted Brown/Taupe
    total_color = "FAF5EE"  # Off-White/Cream for section totals
    alternate_row_color = "F3E3D8"  # Very Light Pink/Peach for alternating rows
    cream_color = "FAF5EE"  # Off-White/Cream
    grand_total_color = "E1CDC0"  # Light Beige/Greige for grand total
    border_color = "F3E3D8"  # Very Light Pink/Peach borders
    
    # Define uniform thin border style
    uniform_border = Border(
        left=Side(style='hair', color=border_color),
        right=Side(style='hair', color=border_color),
        top=Side(style='hair', color=border_color),
        bottom=Side(style='hair', color=border_color)
    )
    
    # Try to load Inter font, fallback to Calibri if not available
    try:
        inter_font = Font(name="Inter", size=10)
        inter_bold = Font(name="Inter", size=11, bold=True)
        inter_header = Font(name="Inter", size=12, bold=True, color=white_color)
    except:
        inter_font = Font(name="Calibri", size=10)
        inter_bold = Font(name="Calibri", size=11, bold=True)
        inter_header = Font(name="Calibri", size=12, bold=True, color=white_color)
    
    # Define professional styles with Inter font
    header_style = NamedStyle(name="header")
    header_style.font = inter_header
    header_style.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_style.border = uniform_border
    
    subheader_style = NamedStyle(name="subheader")
    subheader_style.font = inter_bold
    subheader_style.fill = PatternFill(start_color=subheader_color, end_color=subheader_color, fill_type="solid")
    subheader_style.alignment = Alignment(horizontal="left", vertical="center")
    subheader_style.border = uniform_border
    
    data_style = NamedStyle(name="data")
    data_style.font = inter_font
    data_style.fill = PatternFill(start_color=white_color, end_color=white_color, fill_type="solid")
    data_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, shrink_to_fit=False)
    data_style.border = uniform_border
    
    alternate_data_style = NamedStyle(name="alternate_data")
    alternate_data_style.font = inter_font
    alternate_data_style.fill = PatternFill(start_color=alternate_row_color, end_color=alternate_row_color, fill_type="solid")
    alternate_data_style.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, shrink_to_fit=False)
    alternate_data_style.border = uniform_border
    
    # Right-aligned styles for numeric columns
    data_right_style = NamedStyle(name="data_right")
    data_right_style.font = inter_font
    data_right_style.fill = PatternFill(start_color=white_color, end_color=white_color, fill_type="solid")
    data_right_style.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    data_right_style.border = uniform_border
    
    alternate_data_right_style = NamedStyle(name="alternate_data_right")
    alternate_data_right_style.font = inter_font
    alternate_data_right_style.fill = PatternFill(start_color=alternate_row_color, end_color=alternate_row_color, fill_type="solid")
    alternate_data_right_style.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    alternate_data_right_style.border = uniform_border
    
    total_style = NamedStyle(name="total")
    total_style.font = inter_bold
    total_style.fill = PatternFill(start_color=total_color, end_color=total_color, fill_type="solid")
    total_style.alignment = Alignment(horizontal="right", vertical="center")
    total_style.border = uniform_border
    
    grand_total_style = NamedStyle(name="grand_total")
    grand_total_style.font = inter_bold
    grand_total_style.fill = PatternFill(start_color=grand_total_color, end_color=grand_total_color, fill_type="solid")
    grand_total_style.alignment = Alignment(horizontal="right", vertical="center")
    grand_total_style.border = uniform_border
    
    # Calculate dynamic column widths based on content
    max_section_len = max(len(str(item.get('Category', ''))) for item in items) if items else 10
    max_room_len = max(len(str(item.get('Room', ''))) for item in items) if items else 10
    max_item_len = max(len(str(item.get('ItemName', ''))) for item in items) if items else 15
    max_desc_len = max(len(str(item.get('Description', ''))) for item in items) if items else 30
    
    # Set minimum and maximum widths for better appearance
    section_width = max(25, min(max_section_len + 5, 35))  # Min 25, Max 35
    room_width = max(18, min(max_room_len + 3, 25))        # Min 18, Max 25  
    item_width = max(30, min(max_item_len + 5, 45))        # Min 30, Max 45
    desc_width = max(60, min(max_desc_len // 2, 80))       # Min 60, Max 80 (keep compact width, use dynamic row heights for wrapping)
    
    print(f"[INFO] Enhanced column widths: Section={section_width}, Room={room_width}, ItemName={item_width}, Description={desc_width}")
    
    # Apply column widths with better sizing
    ws.column_dimensions['A'].width = section_width      # Category/Section
    ws.column_dimensions['B'].width = room_width         # Room  
    ws.column_dimensions['C'].width = item_width         # ItemName
    ws.column_dimensions['D'].width = desc_width         # Description
    ws.column_dimensions['E'].width = 12                 # Quantity
    ws.column_dimensions['F'].width = 18                 # UnitCost
    ws.column_dimensions['G'].width = 10                 # Markup
    ws.column_dimensions['H'].width = 12                 # MarkupType
    ws.column_dimensions['I'].width = 15                 # Total
    
    print(f"[INFO] Dynamic column widths: Section={section_width}, Room={room_width}, ItemName={item_width}, Description={desc_width}")

    # Set row height for better spacing
    ws.row_dimensions[1].height = 25  # Header row
    
    # Add headers with professional styling (matching reference image)
    headers = ["Section", "Room", "Item Name", "Description", "Quantity", "Unit Cost", "Markup", "Total", "Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.style = header_style
    
    # Group items by category
    categories = defaultdict(list)
    for item in items:
        category = item.get('Category', '').strip()
        if category:
            categories[category].append(item)
    
    current_row = 2
    overall_subtotal = 0
    
    # Add items by category with professional formatting (matching reference image)
    for category, category_items in categories.items():
        # Add category header with professional styling
        category_cell = ws.cell(row=current_row, column=1, value=category)
        category_cell.style = subheader_style
        # Apply border to all cells in the row
        for col in range(2, 10):
            ws.cell(row=current_row, column=col, value="").style = subheader_style
        current_row += 1
        
        category_total = 0
        
        # Add items in this category with alternating row colors
        for i, item in enumerate(category_items):
            # Clean up Total value
            total_value = item.get('Total', '0')
            if isinstance(total_value, str):
                total_value = total_value.replace(',', '').replace('$', '').strip()
                try:
                    total_value = float(total_value)
                except ValueError:
                    total_value = 0.0
            else:
                total_value = float(total_value) if total_value else 0.0
            
            category_total += total_value
            
            # Choose style based on row number for alternating colors
            row_style = data_style if i % 2 == 0 else alternate_data_style
            row_style_right = data_right_style if i % 2 == 0 else alternate_data_right_style
            
            # Add item row with professional formatting (matching reference image)
            ws.cell(row=current_row, column=1, value="").style = row_style # Section column is empty
            ws.cell(row=current_row, column=2, value=item.get('Room', '')).style = row_style
            ws.cell(row=current_row, column=3, value=item.get('ItemName', '')).style = row_style
            
            # Add description with proper text wrapping and row height
            clean_desc = clean_description_text(item.get('Description', ''))
            desc_cell = ws.cell(row=current_row, column=4, value=clean_desc)
            desc_cell.style = row_style
            
            # Calculate appropriate row height for description wrapping
            if len(clean_desc) > 60:  # If description is longer than column width, increase row height
                # More intelligent line calculation based on actual column width
                estimated_lines = max(2, len(clean_desc) // 60)  # Estimate lines needed based on 60-char column
                row_height = max(20, estimated_lines * 18)  # Min 20, 18 per line for better spacing
                ws.row_dimensions[current_row].height = row_height
                print(f"[DEBUG] Row {current_row}: Description length {len(clean_desc)}, estimated {estimated_lines} lines, height {row_height}")
            
            ws.cell(row=current_row, column=5, value=item.get('Quantity', '')).style = row_style_right
            ws.cell(row=current_row, column=6, value=item.get('UnitCost', '')).style = row_style_right
            ws.cell(row=current_row, column=7, value=item.get('Markup', '')).style = row_style_right
            ws.cell(row=current_row, column=8, value=f"${total_value:,.2f}").style = row_style_right
            ws.cell(row=current_row, column=9, value=item.get('Confidence', '')).style = row_style_right
            current_row += 1
        
        # Add category total with professional styling
        if category_total > 0:
            print(f"[INFO] Category '{category}' total: ${category_total:,.2f}")
            # Add just "Total" label (no section name)
            ws.cell(row=current_row, column=1, value="Total").style = total_style
            # Add empty cells for proper formatting
            for col in range(2, 8):
                ws.cell(row=current_row, column=col, value="").style = total_style
            
            # Add the total with professional formatting
            total_cell = ws.cell(row=current_row, column=8, value=f"${category_total:,.2f}")
            total_cell.style = total_style
            ws.cell(row=current_row, column=9, value="").style = total_style
            current_row += 1
            
            # Add a properly styled blank row after total (fix the extra cell color issue)
            for col in range(1, 10):
                ws.cell(row=current_row, column=col, value="").style = data_style
            current_row += 1
        
        # Add gap between sections
        current_row += 1
        
        overall_subtotal += category_total
    
    # Add professional summary section (matching reference image)
    if overall_subtotal > 0:
        # Add overall subtotal
        ws.cell(row=current_row, column=1, value="Overall Subtotal").style = total_style
        for col in range(2, 8):
            ws.cell(row=current_row, column=col, value="").style = total_style
        ws.cell(row=current_row, column=8, value=f"${overall_subtotal:,.2f}").style = total_style
        ws.cell(row=current_row, column=9, value="").style = total_style
        current_row += 1
        
        # Add general conditions (10%)
        general_conditions = overall_subtotal * 0.10
        ws.cell(row=current_row, column=1, value="General Conditions (10%)").style = total_style
        for col in range(2, 8):
            ws.cell(row=current_row, column=col, value="").style = total_style
        ws.cell(row=current_row, column=8, value=f"${general_conditions:,.2f}").style = total_style
        ws.cell(row=current_row, column=9, value="").style = total_style
        current_row += 1
        
        # Add grand total with professional styling
        grand_total = overall_subtotal + general_conditions
        ws.cell(row=current_row, column=1, value="GRAND TOTAL").style = grand_total_style
        for col in range(2, 8):
            ws.cell(row=current_row, column=col, value="").style = grand_total_style
        grand_total_cell = ws.cell(row=current_row, column=8, value=f"${grand_total:,.2f}")
        grand_total_cell.style = grand_total_style
        ws.cell(row=current_row, column=9, value="").style = grand_total_style
    
    # Save the workbook
    wb.save(output_file)
    print(f"[INFO] Beautiful Excel file created successfully: {output_file}")

def write_final_csv(items, output_file):
    """Write the final cleaned items to CSV."""
    print(f"[INFO] Writing final CSV: {output_file}")
    
    if not items:
        print("[WARNING] No items to write")
        return
    
    # Get all unique categories
    categories = set()
    for item in items:
        category = item.get('Category', '').strip()
        if category:
            categories.add(category)
    
    # Write CSV with sections
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        fieldnames = ['Category', 'Room', 'ItemName', 'Description', 'Quantity', 'UnitCost', 'Markup', 'MarkupType', 'Total', 'Confidence']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        overall_subtotal = 0
        
        for category in sorted(categories):
            # Add category header (only allowed keys)
            writer.writerow({k: (category if k == 'Category' else '') for k in fieldnames})
            
            category_total = 0
            category_items = [item for item in items if item.get('Category', '').strip() == category]
            
            for item in category_items:
                # Filter out any unexpected keys (e.g., 'DeduplicationNotes')
                safe_item = {k: item.get(k, '') for k in fieldnames}
                writer.writerow(safe_item)
                
                # Calculate category total
                total_value = item.get('Total', '0')
                if isinstance(total_value, str):
                    total_value = total_value.replace(',', '').replace('$', '').replace('"', '').replace("'", '').strip()
                    try:
                        total_value = float(total_value)
                    except ValueError:
                        total_value = 0.0
                else:
                    total_value = float(total_value) if total_value else 0.0
                
                category_total += total_value
            
            # Add category total
            if category_total > 0:
                writer.writerow({
                    'Category': '', 'Room': '', 'ItemName': '', 'Description': '', 
                    'Quantity': '', 'UnitCost': '', 'Markup': '', 'MarkupType': '', 
                    'Total': f"{category_total:.2f}", 'Confidence': ''
                })
            
            overall_subtotal += category_total
        
        # Add overall totals
        if overall_subtotal > 0:
            writer.writerow({
                'Category': '', 'Room': '', 'ItemName': '', 'Description': '', 
                'Quantity': '', 'UnitCost': '', 'Markup': '', 'MarkupType': '', 
                'Total': f"{overall_subtotal:.2f}", 'Confidence': ''
            })
            
            general_conditions = overall_subtotal * 0.10
            writer.writerow({
                'Category': '', 'Room': '', 'ItemName': '', 'Description': 'General Conditions (10%)', 
                'Quantity': '', 'UnitCost': '', 'Markup': '', 'MarkupType': '', 
                'Total': f"{general_conditions:.2f}", 'Confidence': ''
            })
            
            grand_total = overall_subtotal + general_conditions
            writer.writerow({
                'Category': '', 'Room': '', 'ItemName': '', 'Description': 'Grand Total', 
                'Quantity': '', 'UnitCost': '', 'Markup': '', 'MarkupType': '', 
                'Total': f"{grand_total:.2f}", 'Confidence': ''
            })
    
    print(f"[INFO] Final CSV written successfully: {output_file}")

def aggregate_chunk_outputs(output_dir):
    """Aggregate outputs from chunked estimation with enhanced unique item capture."""
    print(f"[INFO] Aggregating chunk outputs from: {output_dir}")
    
    # Find all chunk output text files
    chunk_files = []
    for file in os.listdir(output_dir):
        if file.startswith('estimate_output_chunk_') and file.endswith('.txt'):
            chunk_files.append(os.path.join(output_dir, file))
    
    if not chunk_files:
        print("[ERROR] No chunk output files found in output directory")
        return None
    
    chunk_files.sort()  # Sort by chunk number
    print(f"[INFO] Found {len(chunk_files)} chunk files")
    
    all_items = []
    seen_items = set()  # Track unique items to avoid duplicates
    
    for chunk_file in chunk_files:
        print(f"[INFO] Processing {chunk_file}")
        try:
            with open(chunk_file, 'r', encoding='utf-8') as f:
                content = f.read()
                # Find JSON start and end
                start = content.find('{')
                end = content.rfind('}') + 1
                if start != -1 and end != 0:
                    json_str = content[start:end]
                    data = json.loads(json_str)
                    
                    if 'sections' in data:
                        for section in data['sections']:
                            section_name = section.get('name', '')
                            for item in section.get('items', []):
                                csv_item = {
                                    'Category': section_name,
                                    'Room': item.get('room', ''),
                                    'ItemName': item.get('scope_item', ''),
                                    'Description': clean_description_text(item.get('description', '')),
                                    'Quantity': item.get('quantity', ''),
                                    'UnitCost': item.get('unit_cost', ''),
                                    'Markup': item.get('markup', ''),
                                    'MarkupType': '%',
                                    'Total': item.get('subtotal', ''),
                                    'Confidence': item.get('confidence_score', '')
                                }
                                
                                # Create unique identifier for deduplication
                                unique_id = f"{csv_item['Category']}|{csv_item['Room']}|{csv_item['ItemName']}"
                                
                                # Only add if we haven't seen this exact item before
                                if unique_id not in seen_items:
                                    seen_items.add(unique_id)
                                    all_items.append(csv_item)
                                else:
                                    print(f"[INFO] Skipping duplicate item: {csv_item['ItemName']} in {csv_item['Room']}")
        except Exception as e:
            print(f"[ERROR] Error processing {chunk_file}: {e}")
    
    print(f"[INFO] Aggregated {len(all_items)} unique items from all chunks")
    
    if not all_items:
        print("[ERROR] No items found in chunks")
        return None
    
    # Write aggregated CSV
    output_csv = os.path.join(output_dir, 'aggregated_chunked_estimate.csv')
    with open(output_csv, 'w', newline='', encoding='utf-8') as f:
        fieldnames = ['Category', 'Room', 'ItemName', 'Description', 'Quantity', 'UnitCost', 'Markup', 'MarkupType', 'Total', 'Confidence']
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_items)
    
    print(f"[INFO] Aggregated CSV written: {output_csv}")
    
    # Perform comprehensive cleanup
    cleaned_items = comprehensive_cleanup(all_items)
    print(f"[INFO] After comprehensive cleanup: {len(cleaned_items)} items")
    
    # Write final CSV
    final_csv = os.path.join(output_dir, 'comprehensive_clean_estimate.csv')
    write_final_csv(cleaned_items, final_csv)
    print(f"[INFO] Final CSV written: {final_csv}")
    
    # Create Excel file
    final_excel = os.path.join(output_dir, 'final_renovation_estimate.xlsx')
    create_excel_file(cleaned_items, final_excel)
    print(f"[INFO] Final Excel written: {final_excel}")
    
    return cleaned_items

def calculate_formula_value(formula_str):
    """Calculate actual value from formula string in subtotal field."""
    if not formula_str or formula_str == "":
        return 0.0
    
    try:
        # Remove any extra spaces and parentheses
        formula = formula_str.strip()
        
        # Handle common patterns
        if "*" in formula and "1.75" in formula:
            # Extract the calculation part before the markup
            parts = formula.split("* 1.75")
            if len(parts) > 0:
                base_calc = parts[0].strip()
                # Evaluate the base calculation
                result = eval(base_calc) * 1.75
                return round(result, 2)
        
        # Handle countertop formula: (1500 + (99 * 97) + 2000) * 1.75
        if "1500" in formula and "97" in formula and "2000" in formula:
            # Extract the calculation inside parentheses
            start = formula.find("(")
            end = formula.find(")")
            if start != -1 and end != -1:
                inner_calc = formula[start+1:end]
                result = eval(inner_calc) * 1.75
                return round(result, 2)
        
        # Handle simple calculations like "321 * 6 * 1.75"
        if "*" in formula:
            result = eval(formula)
            return round(result, 2)
        
        # Try to convert to float directly
        return float(formula)
        
    except Exception as e:
        print(f"[WARNING] Could not calculate formula '{formula_str}': {e}")
        return 0.0

def fix_total_values(items):
    """Fix total values by calculating formulas and ensuring proper formatting."""
    print(f"[INFO] Fixing total values for {len(items)} items")
    
    for item in items:
        total_value = item.get('Total', '')
        
        if isinstance(total_value, str):
            # Check if it's a formula
            if any(op in total_value for op in ['*', '+', '-', '/', '(']):
                calculated_value = calculate_formula_value(total_value)
                item['Total'] = str(calculated_value)
                print(f"[INFO] Calculated formula '{total_value}' = {calculated_value}")
            elif total_value.strip() == "":
                # Empty total, try to calculate from quantity and unit cost
                quantity = item.get('Quantity', '0')
                unit_cost = item.get('UnitCost', '0')
                markup = float(item.get('Markup', '0.75'))
                
                try:
                    # Extract numeric value from quantity
                    qty_num = 0
                    if 'SF' in quantity:
                        qty_num = float(quantity.replace('SF', '').strip())
                    elif 'LF' in quantity:
                        qty_num = float(quantity.replace('LF', '').strip())
                    elif 'UNIT' in quantity:
                        qty_num = float(quantity.replace('UNIT', '').strip())
                    else:
                        qty_num = float(quantity)
                    
                    # Extract numeric value from unit cost
                    unit_num = 0
                    if isinstance(unit_cost, str):
                        # Extract first number from unit cost
                        import re
                        numbers = re.findall(r'\d+', unit_cost)
                        if numbers:
                            unit_num = float(numbers[0])
                    
                    calculated_total = qty_num * unit_num * (1 + markup)
                    item['Total'] = str(round(calculated_total, 2))
                    print(f"[INFO] Calculated total for {item.get('ItemName', '')}: {calculated_total}")
                    
                except Exception as e:
                    print(f"[WARNING] Could not calculate total for {item.get('ItemName', '')}: {e}")
                    item['Total'] = '0'
    
    return items

def is_commercial_cleaning_duplicate(item1, item2):
    """Check if two items are commercial cleaning duplicates."""
    name1 = item1.get('ItemName', '').lower()
    name2 = item2.get('ItemName', '').lower()
    room1 = item1.get('Room', '').lower()
    room2 = item2.get('Room', '').lower()
    
    # Check if both are commercial cleaning
    if 'cleaning' in name1 and 'cleaning' in name2:
        # If same room, it's a duplicate
        if room1 == room2:
            return True
        # If one is "entire apartment" and other is specific room, keep the specific one
        if 'entire' in room1 and 'entire' not in room2:
            return True
        if 'entire' in room2 and 'entire' not in room1:
            return True
    
    return False

def is_flooring_duplicate(item1, item2):
    """Check if two items are flooring duplicates."""
    name1 = normalize_item_name(item1.get('ItemName', ''))
    name2 = normalize_item_name(item2.get('ItemName', ''))
    room1 = normalize_room_name(item1.get('Room', ''))
    room2 = normalize_room_name(item2.get('Room', ''))
    
    # If rooms are different, they're not duplicates
    if room1 != room2:
        return False
    
    # Check if both are flooring related (more comprehensive keywords)
    flooring_keywords = ['floor', 'flooring', 'hardwood', 'laminate', 'tile floor', 'tile flooring', 'subfloor', 'underlayment']
    desc1 = normalize_item_name(item1.get('Description', ''))
    desc2 = normalize_item_name(item2.get('Description', ''))
    
    # Check both name and description for flooring keywords
    combined1 = f"{name1} {desc1}".lower()
    combined2 = f"{name2} {desc2}".lower()
    
    if any(keyword in combined1 for keyword in flooring_keywords) and any(keyword in combined2 for keyword in flooring_keywords):
        return True
    
    return False

def is_waterproofing_duplicate(item1, item2):
    """Check if two items are waterproofing duplicates."""
    name1 = item1.get('ItemName', '').lower()
    name2 = item2.get('ItemName', '').lower()
    room1 = item1.get('Room', '').lower()
    room2 = item2.get('Room', '').lower()
    
    # Check if both are waterproofing related
    waterproofing_keywords = ['waterproof', 'waterproofing', 'membrane']
    if any(keyword in name1 for keyword in waterproofing_keywords) and any(keyword in name2 for keyword in waterproofing_keywords):
        if room1 == room2:
            return True
    
    return False

def is_walls_ceilings_duplicate(item1, item2):
    """Check if two items are walls/ceilings duplicates."""
    name1 = item1.get('ItemName', '').lower()
    name2 = item2.get('ItemName', '').lower()
    room1 = item1.get('Room', '').lower()
    room2 = item2.get('Room', '').lower()
    
    # Check if both are walls/ceilings related
    walls_keywords = ['wall', 'ceiling', 'drywall', 'plaster']
    if any(keyword in name1 for keyword in walls_keywords) and any(keyword in name2 for keyword in walls_keywords):
        if room1 == room2:
            return True
    
    return False

def is_backsplash_duplicate(item1, item2):
    """Check if two items are backsplash duplicates."""
    name1 = item1.get('ItemName', '').lower()
    name2 = item2.get('ItemName', '').lower()
    room1 = item1.get('Room', '').lower()
    room2 = item2.get('Room', '').lower()
    
    # Check if both are backsplash related
    if 'backsplash' in name1 and 'backsplash' in name2:
        if room1 == room2:
            return True
    
    return False

def is_trim_duplicate(item1, item2):
    """Check if two items are trim duplicates."""
    name1 = item1.get('ItemName', '').lower()
    name2 = item2.get('ItemName', '').lower()
    room1 = item1.get('Room', '').lower()
    room2 = item2.get('Room', '').lower()
    
    # Check if both are trim related
    trim_keywords = ['trim', 'baseboard', 'molding']
    if any(keyword in name1 for keyword in trim_keywords) and any(keyword in name2 for keyword in trim_keywords):
        if room1 == room2:
            return True
    
    return False

def is_doors_duplicate(item1, item2):
    """Check if two items are doors duplicates."""
    name1 = item1.get('ItemName', '').lower()
    name2 = item2.get('ItemName', '').lower()
    room1 = item1.get('Room', '').lower()
    room2 = item2.get('Room', '').lower()
    
    # Check if both are doors related
    door_keywords = ['door', 'entry', 'interior door', 'exterior door']
    if any(keyword in name1 for keyword in door_keywords) and any(keyword in name2 for keyword in door_keywords):
        if room1 == room2:
            return True
    
    return False

def is_cabinetry_duplicate(item1, item2):
    """Check if two items are cabinetry duplicates."""
    name1 = item1.get('ItemName', '').lower()
    name2 = item2.get('ItemName', '').lower()
    room1 = item1.get('Room', '').lower()
    room2 = item2.get('Room', '').lower()
    
    # If rooms are different, they're not duplicates
    if room1 != room2:
        return False
    
    # Check if both are cabinetry related
    cabinet_keywords = ['cabinet', 'cabinetry', 'kitchen cabinet', 'bathroom cabinet', 'closet']
    if any(keyword in name1 for keyword in cabinet_keywords) and any(keyword in name2 for keyword in cabinet_keywords):
        return True
    
    return False

def merge_cabinetry_categories(items):
    """Merge 'Cabinetry' and 'Cabinetry & Storage' categories into one."""
    print(f"[INFO] Merging cabinetry categories in {len(items)} items")
    
    # Group items by category
    categories = defaultdict(list)
    for item in items:
        category = item.get('Category', '').strip()
        if category:
            categories[category].append(item)
    
    merged_items = []
    
    for category, category_items in categories.items():
        if category in ['Cabinetry', 'Cabinetry & Storage']:
            # Change category to 'Cabinetry & Storage' for all cabinetry items
            for item in category_items:
                item['Category'] = 'Cabinetry & Storage'
            merged_items.extend(category_items)
        else:
            # Keep other categories as is
            merged_items.extend(category_items)
    
    print(f"[INFO] Merged cabinetry categories: {len(items)} -> {len(merged_items)} items")
    return merged_items

def remove_smaller_rooms_when_full_apartment_exists(items):
    """Remove smaller room items when apartment-level scope exists, based on reference CSV patterns."""
    print(f"[INFO] Checking for full apartment vs smaller room conflicts in {len(items)} items")
    
    # Based on reference CSV analysis, these categories are typically apartment-level
    apartment_level_categories = {
        'Demolition',           # Always apartment-level in references
        'Electrical',           # Usually apartment-level in references  
        'Flooring',            # Usually apartment-level in references
        'Trims',               # Always apartment-level in references
        'Painting & Wall Coverings',  # Always apartment-level in references
        'Heating and Cooling', # Always apartment-level in references
        'Windows',             # Always apartment-level in references
        'General Requirements', # Always apartment-level in references
        'Doors'                # Can be apartment-level
    }
    
    # These categories are typically room-specific
    room_specific_categories = {
        'Plumbing',            # Room-specific in references
        'Tile',                # Room-specific in references  
        'Waterproofing',       # Room-specific in references
        'Cabinetry',           # Kitchen-specific in references
        'Countertops',         # Kitchen-specific in references
        'Backsplash',          # Kitchen-specific in references
        'Accessories'          # Bathroom-specific in references
    }
    
    # Group items by category
    categories = defaultdict(list)
    for item in items:
        cat = item.get("Category", "").strip()
        categories[cat].append(item)

    cleaned = []
    
    for cat, cat_items in categories.items():
        if cat in apartment_level_categories:
            # For apartment-level categories, prefer general/apartment scope over specific rooms
            apartment_items = []
            room_items = []
            
            for item in cat_items:
                room = item.get('Room', '').lower().strip()
                if ('general' in room or 'apartment' in room or 'entire' in room or 
                    'overall' in room or 'building' in room or room == ''):
                    apartment_items.append(item)
                else:
                    room_items.append(item)
            
            if apartment_items:
                print(f"[INFO] Found apartment-level scope in '{cat}', removing {len(room_items)} smaller items")
                cleaned.extend(apartment_items)
            else:
                # No apartment-level scope, keep all room items
                cleaned.extend(cat_items)
                
        elif cat in room_specific_categories:
            # For room-specific categories, keep all items (they should be room-specific)
            cleaned.extend(cat_items)
        else:
            # For other categories, use the existing logic
            full_items = []
            other_items = []
            
            for item in cat_items:
                room = item.get('Room', '').lower().strip()
                if ('entire' in room or 'full' in room or 'whole' in room or 
                    'apartment' in room or 'general' in room or 'all rooms' in room):
                    full_items.append(item)
                else:
                    other_items.append(item)
            
            if full_items:
                print(f"[INFO] Found apartment-level scope in '{cat}', removing {len(other_items)} smaller items")
                cleaned.extend(full_items)
            else:
                cleaned.extend(cat_items)
    
    print(f"[INFO] After removing smaller rooms: {len(items)} -> {len(cleaned)} items")
    return cleaned

def aggregate_chunk_outputs(run_dir):
    """Aggregate text outputs from chunk processing into CSV format."""
    print(f"[INFO] Aggregating chunk outputs from: {run_dir}")
    
    # Find all estimate output files
    output_files = []
    for file in os.listdir(run_dir):
        if file.startswith('estimate_output_chunk_') and file.endswith('.txt'):
            output_files.append(os.path.join(run_dir, file))
    
    if not output_files:
        print("[ERROR] No estimate output files found")
        return None
    
    # Sort files by chunk number
    output_files.sort(key=lambda x: int(x.split('_chunk_')[1].split('.')[0]))
    
    all_items = []
    
    for file_path in output_files:
        print(f"[INFO] Processing: {os.path.basename(file_path)}")
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Parse the content to extract items
            items = parse_estimation_output(content)
            if items:
                all_items.extend(items)
                print(f"[INFO] Found {len(items)} items in {os.path.basename(file_path)}")
            else:
                print(f"[WARNING] No items found in {os.path.basename(file_path)}")
                
        except Exception as e:
            print(f"[ERROR] Failed to process {file_path}: {e}")
    
    if all_items:
        # Write aggregated CSV
        output_csv = os.path.join(run_dir, 'comprehensive_clean_estimate.csv')
        write_final_csv(all_items, output_csv)
        print(f"[SUCCESS] Aggregated {len(all_items)} items to {output_csv}")
        return all_items
    else:
        print("[ERROR] No items aggregated from any chunks")
        return None

def parse_estimation_output(content):
    """Parse estimation output text and extract structured items from JSON format."""
    items = []
    
    try:
        # Try to find JSON content in the response
        import json
        import re
        
        # Look for JSON content between ```json and ``` markers
        json_match = re.search(r'```json\s*(.*?)\s*```', content, re.DOTALL)
        if json_match:
            json_content = json_match.group(1)
            data = json.loads(json_content)
            
            # Extract items from the JSON structure
            if 'sections' in data:
                for section in data['sections']:
                    section_name = section.get('name', '')
                    if 'items' in section:
                        for item in section['items']:
                            # Convert the new format to the expected format
                            converted_item = {
                                'Category': section_name,
                                'ItemName': item.get('scope_item', ''),
                                'Description': clean_description_text(item.get('description', '')),
                                'Room': item.get('room', ''),
                                'Quantity': item.get('quantity', ''),
                                # Preserve unit cost text for downstream numeric extraction
                                'UnitCost': item.get('unit_cost', ''),
                                # Map subtotal to Total so cleanup can calculate/roll-up
                                'Total': item.get('subtotal', ''),
                                'Markup': item.get('markup', ''),
                                'Confidence': item.get('confidence_score', '')
                            }
                            items.append(converted_item)
            
            print(f"[INFO] Successfully parsed JSON with {len(items)} items")
            return items
            
    except Exception as e:
        print(f"[WARNING] JSON parsing failed: {e}")
    
    # Fallback to old text parsing if JSON fails
    print("[INFO] Falling back to text parsing...")
    lines = content.split('\n')
    current_item = {}
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Look for patterns that indicate item boundaries
        if 'Category:' in line or 'Item:' in line or 'Description:' in line:
            if current_item and 'Category' in current_item and 'ItemName' in current_item:
                items.append(current_item)
                current_item = {}
            
            # Extract category and item name
            if 'Category:' in line:
                category = line.split('Category:')[1].strip()
                current_item['Category'] = category
            elif 'Item:' in line:
                item_name = line.split('Item:')[1].strip()
                current_item['ItemName'] = item_name
            elif 'Description:' in line:
                description = line.split('Description:')[1].strip()
                current_item['Description'] = clean_description_text(description)
                
        # Look for other fields
        elif 'Room:' in line:
            room = line.split('Room:')[1].strip()
            current_item['Room'] = room
        elif 'Quantity:' in line:
            quantity = line.split('Quantity:')[1].strip()
            current_item['Quantity'] = quantity
        elif 'Unit:' in line:
            unit = line.split('Unit:')[1].strip()
            current_item['Unit'] = unit
        elif 'Price:' in line:
            price = line.split('Price:')[1].strip()
            current_item['Price'] = price
    
    # Add the last item if it exists
    if current_item and 'Category' in current_item and 'ItemName' in current_item:
        items.append(current_item)
    
    return items

def merge_overlapping_items(items):
    """Merge overlapping items to create comprehensive, non-duplicate estimates."""
    print(f"[INFO] Merging overlapping items from {len(items)} items")
    
    # Group by category and room
    grouped_items = defaultdict(lambda: defaultdict(list))
    for item in items:
        category = item.get('Category', '').strip()
        room = normalize_room_name(item.get('Room', ''))
        if category and room:
            grouped_items[category][room].append(item)
    
    merged_items = []
    
    for category, rooms in grouped_items.items():
        for room, room_items in rooms.items():
            if len(room_items) == 1:
                # Single item, keep as-is
                merged_items.append(room_items[0])
            else:
                # Multiple items in same room, try to merge
                merged = merge_room_items(room_items, category, room)
                merged_items.extend(merged)
    
    print(f"[INFO] After merging: {len(merged_items)} items")
    return merged_items

def merge_room_items(room_items, category, room):
    """Merge multiple items in the same room to eliminate duplicates."""
    if len(room_items) == 1:
        return room_items
    
    # Sort by confidence and specificity
    room_items.sort(key=lambda x: (float(x.get('Confidence', 0)), -len(x.get('Description', ''))), reverse=True)
    
    merged_items = []
    processed_work_types = set()
    
    for item in room_items:
        work_type = identify_work_type(item)
        
        if work_type in processed_work_types:
            # Skip if we already have this type of work in this room
            print(f"[INFO] Skipping duplicate work type '{work_type}' in {room}: {item.get('ItemName', '')}")
            continue
        
        # Check if this item is a subset of an existing item in the same room
        is_subset = False
        for existing_item in merged_items:
            if is_work_subset(item, existing_item):
                print(f"[INFO] Merging subset work in same room: {item.get('ItemName', '')} into {existing_item.get('ItemName', '')} in {room}")
                is_subset = True
                break
        
        if not is_subset:
            merged_items.append(item)
            processed_work_types.add(work_type)
            print(f"[INFO] Keeping unique work: {item.get('ItemName', '')} in {room}")
    
    return merged_items

def identify_work_type(item):
    """Identify the type of work being performed."""
    name = normalize_item_name(item.get('ItemName', ''))
    desc = normalize_item_name(item.get('Description', ''))
    combined = f"{name} {desc}".lower()
    
    work_types = {
        'demolition': ['demolition', 'demo', 'gut', 'remove', 'tear out', 'strip'],
        'electrical': ['electrical', 'wiring', 'outlet', 'switch', 'light', 'panel', 'rewiring'],
        'plumbing': ['plumbing', 'plumb', 'sink', 'toilet', 'shower', 'tub', 'fixture'],
        'cabinetry': ['cabinet', 'cabinetry', 'storage', 'shelf', 'drawer'],
        'countertop': ['countertop', 'counter', 'quartz', 'marble', 'granite', 'surface'],
        'tile': ['tile', 'tiling', 'ceramic', 'porcelain', 'grout'],
        'flooring': ['flooring', 'floor', 'hardwood', 'laminate', 'vinyl'],
        'painting': ['paint', 'painting', 'primer', 'coat', 'finish'],
        'backsplash': ['backsplash', 'back splash', 'wall tile'],
        'trim': ['trim', 'baseboard', 'crown', 'molding', 'moulding'],
        'doors': ['door', 'frame', 'jamb', 'hinge'],
        'waterproofing': ['waterproof', 'water proof', 'moisture', 'seal'],
        'cleaning': ['clean', 'cleaning', 'post construction', 'final clean'],
        'appliances': ['appliance', 'oven', 'range', 'microwave', 'dishwasher', 'refrigerator']
    }
    
    for work_type, keywords in work_types.items():
        if any(keyword in combined for keyword in keywords):
            return work_type
    
    return 'general'

def is_work_subset(item1, item2):
    """Check if item1 is a subset of item2's work scope."""
    name1 = normalize_item_name(item1.get('ItemName', ''))
    name2 = normalize_item_name(item2.get('ItemName', ''))
    desc1 = normalize_item_name(item1.get('Description', ''))
    desc2 = normalize_item_name(item2.get('Description', ''))
    
    # Check if one item's description encompasses the other
    combined1 = f"{name1} {desc1}".lower()
    combined2 = f"{name2} {desc2}".lower()
    
    # Check for subset relationships
    if 'full gut' in combined2 and any(word in combined1 for word in ['demolition', 'remove', 'gut']):
        return True
    
    if 'complete' in combined2 and any(word in combined1 for word in ['rewiring', 'electrical', 'plumbing']):
        return True
    
    if 'installation' in combined2 and any(word in combined1 for word in ['install', 'put in', 'set up']):
        return True
    
    return False

def remove_cross_category_duplicates(items):
    """Remove items that represent the same work but are categorized in different sections."""
    print(f"[INFO] Removing cross-category duplicates from {len(items)} items")
    
    # Group items by room and work type
    room_work_groups = defaultdict(lambda: defaultdict(list))
    
    for item in items:
        room = normalize_room_name(item.get('Room', ''))
        work_type = identify_work_type(item)
        if room and work_type:
            room_work_groups[room][work_type].append(item)
    
    cleaned_items = []
    
    for room, work_types in room_work_groups.items():
        for work_type, type_items in work_types.items():
            if len(type_items) == 1:
                # Single item for this work type in this room, keep it
                cleaned_items.append(type_items[0])
            else:
                # Multiple items for same work type in same room, keep the best one
                best_item = select_best_item(type_items)
                cleaned_items.append(best_item)
                
                # Log what was removed
                removed_items = [item for item in type_items if item != best_item]
                for removed in removed_items:
                    print(f"[INFO] Removed cross-category duplicate in same room: {removed.get('ItemName', '')} in {removed.get('Category', '')} for {room} {work_type}")
    
    print(f"[INFO] Cross-category deduplication complete: {len(items)} -> {len(cleaned_items)} items")
    return cleaned_items

def select_best_item(items):
    """Select the best item from a list of similar items based on multiple criteria."""
    if not items:
        return None
    
    # Score each item based on multiple criteria
    scored_items = []
    for item in items:
        score = 0
        
        # Higher confidence = higher score
        try:
            confidence = float(item.get('Confidence', 0))
            score += confidence * 10
        except:
            pass
        
        # More specific description = higher score
        desc_length = len(item.get('Description', ''))
        score += desc_length * 0.1
        
        # Better category = higher score (some categories are more appropriate for certain work)
        category = item.get('Category', '').lower()
        work_type = identify_work_type(item)
        
        # Category appropriateness scoring
        category_scores = {
            'demolition': {'demolition': 100, 'electrical': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0},
            'electrical': {'electrical': 100, 'demolition': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0},
            'plumbing': {'plumbing': 100, 'demolition': 0, 'electrical': 0, 'tile': 0, 'flooring': 0},
            'tile': {'tile': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'flooring': 50},
            'flooring': {'flooring': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 50},
            'painting': {'painting': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0},
            'cabinetry': {'cabinetry': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0},
            'countertop': {'countertop': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0},
            'backsplash': {'backsplash': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 50, 'flooring': 0},
            'trim': {'trim': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0},
            'doors': {'doors': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0},
            'waterproofing': {'waterproofing': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0},
            'cleaning': {'cleaning': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0},
            'appliances': {'appliances': 100, 'demolition': 0, 'electrical': 0, 'plumbing': 0, 'tile': 0, 'flooring': 0}
        }
        
        if work_type in category_scores and category in category_scores[work_type]:
            score += category_scores[work_type][category]
        
        scored_items.append((item, score))
    
    # Sort by score (highest first) and return the best
    scored_items.sort(key=lambda x: x[1], reverse=True)
    return scored_items[0][0]

def main():
    # Find the latest output directory from chunked_outputs
    chunked_outputs_dir = 'chunked_outputs'
    if not os.path.exists(chunked_outputs_dir):
        print(f"[ERROR] {chunked_outputs_dir} directory not found")
        return
    
    # Get all run directories
    run_dirs = []
    for item in os.listdir(chunked_outputs_dir):
        item_path = os.path.join(chunked_outputs_dir, item)
        if item.startswith('run_') and os.path.isdir(item_path):
            run_dirs.append(item_path)
    
    if not run_dirs:
        print("[ERROR] No run directories found in chunked_outputs")
        return
    
    # Get the most recent run directory
    latest_dir = max(run_dirs, key=os.path.getctime)
    print(f"[INFO] Processing directory: {latest_dir}")
    
    # Check if we need to aggregate chunk outputs first
    csv_files = []
    for file in os.listdir(latest_dir):
        if (file.startswith('comprehensive_clean_estimate_') or file == 'comprehensive_clean_estimate.csv') and file.endswith('.csv'):
            csv_files.append(os.path.join(latest_dir, file))
    
    if not csv_files:
        print("[INFO] No CSV files found, attempting to aggregate chunk outputs...")
        aggregated_items = aggregate_chunk_outputs(latest_dir)
        if aggregated_items:
            # Now look for the CSV file again
            for file in os.listdir(latest_dir):
                if (file.startswith('comprehensive_clean_estimate_') or file == 'comprehensive_clean_estimate.csv') and file.endswith('.csv'):
                    csv_files.append(os.path.join(latest_dir, file))
    
    if not csv_files:
        print("[ERROR] No comprehensive clean CSV files found after aggregation")
        return
    
    # Get the most recent CSV file
    latest_csv = max(csv_files, key=os.path.getctime)
    print(f"[INFO] Processing CSV file: {latest_csv}")
    
    # Read the CSV file
    items = read_csv_items(latest_csv)
    print(f"[INFO] Read {len(items)} items from CSV")
    
    # Perform comprehensive cleanup
    cleaned_items = comprehensive_cleanup(items)
    print(f"[INFO] After comprehensive cleanup: {len(cleaned_items)} items")
    
    # Write final CSV
    output_csv = latest_csv.replace('.csv', '_final.csv')
    write_final_csv(cleaned_items, output_csv)
    print(f"[INFO] Final CSV written: {output_csv}")
    
    # Create Excel file
    output_excel = latest_csv.replace('.csv', '_final.xlsx')
    create_excel_file(cleaned_items, output_excel)
    print(f"[INFO] Final Excel written: {output_excel}")
    
    print("[INFO] Comprehensive cleanup completed successfully!")

if __name__ == "__main__":
    main() 