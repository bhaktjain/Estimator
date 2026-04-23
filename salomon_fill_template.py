from __future__ import annotations

from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook


def main() -> None:
    template_path = Path("/Users/bhakt/Downloads/SALOMON_NYC_Bid_Estimate_EN_190126.xlsx")
    out_dir = Path("/Users/bhakt/Desktop/custom gpt/outputs/salomon_nyc_20260120")
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / "SALOMON_NYC_Bid_Estimate_EN_190126_FILLED.xlsx"

    wb = load_workbook(template_path, data_only=False)

    # ---------------------------------------------------------------------
    # Quantities derived from provided PDFs (Survey + ESQ) + explicit assumptions
    # - ESQ shows: Kitchen floor = 70 m²; and a separate 64.4 m² callout (treated as Design Studio)
    # - Survey plan is not a clean area schedule; we use 2,763 SF as the overall takeoff area
    #   per the drawing set context (matches prior plan set patterns).
    # ---------------------------------------------------------------------
    total_area_sqm = 256.6  # 2,763 SF ≈ 256.6 m²
    kitchen_terrazzo_sqm = 70.0
    design_studio_carpet_ege_sqm = 64.4
    terrazzo_other_sqm = 30.0
    oak_vip_sqm = 25.0
    carpet_interface_sqm = max(
        0.0,
        total_area_sqm - (kitchen_terrazzo_sqm + terrazzo_other_sqm + oak_vip_sqm + design_studio_carpet_ege_sqm),
    )

    # Assumptions (to be verified)
    solid_partitions_sqm = 120.0
    glazed_partitions_lf = 120.0
    doors_ea = 18.0
    demo_floor_finishes_sqm = total_area_sqm
    haul_off_tons = 20.0
    paint_walls_ceilings_sqm = total_area_sqm * 2.5
    cork_wall_sqm = 20.0
    photowall_sqm = 12.0
    ceilings_repair_sqm = total_area_sqm
    ceilings_paint_sqm = total_area_sqm
    wall_cladding_sqm = 80.0
    entry_lockers_ea = 30.0
    showroom_bench_ea = 4.0
    water_fountain_cab_ea = 1.0
    tv_cabinets_ea = 4.0
    clareo_luminaires_ea = 120.0
    tal_lighting_ea = 40.0
    decorative_lighting_ea = 25.0
    kitchen_cabinetry_lf = 40.0
    stainless_counter_lf = 35.0
    grohe_faucet_ea = 2.0
    sanitary_tile_sqm = 55.0
    sanitary_fixtures_ea = 20.0
    workstations_qty = 30.0
    furniture_items_ea = 120.0
    window_film_sqm = 111.5
    logos_ea = 8.0
    racks_ea = 10.0

    # Map (sheet,row) -> quantity/unit price. Columns: D=Quantity, E=Unit Price, F=Amount
    quantities: Dict[Tuple[str, int], float] = {}
    unit_prices: Dict[Tuple[str, int], float] = {}

    def q(sheet: str, row: int, val: float) -> None:
        quantities[(sheet, row)] = val

    def p(sheet: str, row: int, val: float) -> None:
        unit_prices[(sheet, row)] = val

    # Trade 01 - Demolition
    q("Trade 01 - Demolition", 2, 120.0)  # 1.1 solid partitions (sq m)
    p("Trade 01 - Demolition", 2, 30.0)
    q("Trade 01 - Demolition", 3, glazed_partitions_lf)  # 1.2 glazed partitions (lf)
    p("Trade 01 - Demolition", 3, 60.0)
    q("Trade 01 - Demolition", 4, demo_floor_finishes_sqm)  # 1.3 floor finishes (sq m)
    p("Trade 01 - Demolition", 4, 35.0)
    p("Trade 01 - Demolition", 5, 5000.0)  # 1.4 protection (lump qty=1 in template)
    q("Trade 01 - Demolition", 6, haul_off_tons)  # 1.5 haul-off (t)
    p("Trade 01 - Demolition", 6, 450.0)

    # Trade 02 - Partitions
    q("Trade 02 - Partitions", 2, solid_partitions_sqm)
    p("Trade 02 - Partitions", 2, 180.0)
    q("Trade 02 - Partitions", 3, glazed_partitions_lf)
    p("Trade 02 - Partitions", 3, 650.0)
    q("Trade 02 - Partitions", 4, doors_ea)
    p("Trade 02 - Partitions", 4, 1200.0)
    p("Trade 02 - Partitions", 5, 8000.0)  # sealing lump

    # Trade 03 - Flooring
    q("Trade 03 - Flooring", 2, design_studio_carpet_ege_sqm)
    p("Trade 03 - Flooring", 2, 220.0)
    q("Trade 03 - Flooring", 3, carpet_interface_sqm)
    p("Trade 03 - Flooring", 3, 200.0)
    q("Trade 03 - Flooring", 4, kitchen_terrazzo_sqm + terrazzo_other_sqm)
    p("Trade 03 - Flooring", 4, 350.0)
    q("Trade 03 - Flooring", 5, oak_vip_sqm)
    p("Trade 03 - Flooring", 5, 400.0)
    q("Trade 03 - Flooring", 6, 450.0)  # skirting LF
    p("Trade 03 - Flooring", 6, 35.0)

    # Trade 04 - Ceilings
    q("Trade 04 - Ceilings", 2, ceilings_repair_sqm)
    p("Trade 04 - Ceilings", 2, 60.0)
    q("Trade 04 - Ceilings", 3, ceilings_paint_sqm)
    p("Trade 04 - Ceilings", 3, 45.0)

    # Trade 05 - Custom Millwork
    q("Trade 05 - Custom Millwork", 2, wall_cladding_sqm)
    p("Trade 05 - Custom Millwork", 2, 450.0)
    q("Trade 05 - Custom Millwork", 3, entry_lockers_ea)
    p("Trade 05 - Custom Millwork", 3, 800.0)
    q("Trade 05 - Custom Millwork", 4, showroom_bench_ea)
    p("Trade 05 - Custom Millwork", 4, 3500.0)
    q("Trade 05 - Custom Millwork", 5, water_fountain_cab_ea)
    p("Trade 05 - Custom Millwork", 5, 4500.0)
    q("Trade 05 - Custom Millwork", 6, tv_cabinets_ea)
    p("Trade 05 - Custom Millwork", 6, 3000.0)
    p("Trade 05 - Custom Millwork", 7, 45000.0)  # bar/backbar lump

    # Trade 06 - Paint & Walls
    q("Trade 06 - Paint & Walls", 2, paint_walls_ceilings_sqm)
    p("Trade 06 - Paint & Walls", 2, 25.0)
    q("Trade 06 - Paint & Walls", 3, cork_wall_sqm)
    p("Trade 06 - Paint & Walls", 3, 150.0)
    q("Trade 06 - Paint & Walls", 4, photowall_sqm)
    p("Trade 06 - Paint & Walls", 4, 120.0)

    # Trade 07 - Electrical & Lightin
    p("Trade 07 - Electrical & Lightin", 2, 180000.0)  # power/data lump
    q("Trade 07 - Electrical & Lightin", 3, clareo_luminaires_ea)
    p("Trade 07 - Electrical & Lightin", 3, 650.0)
    q("Trade 07 - Electrical & Lightin", 4, tal_lighting_ea)
    p("Trade 07 - Electrical & Lightin", 4, 800.0)
    q("Trade 07 - Electrical & Lightin", 5, decorative_lighting_ea)
    p("Trade 07 - Electrical & Lightin", 5, 1200.0)
    p("Trade 07 - Electrical & Lightin", 6, 15000.0)  # install/testing lump

    # Trade 08 - Kitchen & Bar
    q("Trade 08 - Kitchen & Bar", 2, kitchen_cabinetry_lf)
    p("Trade 08 - Kitchen & Bar", 2, 1800.0)
    q("Trade 08 - Kitchen & Bar", 3, stainless_counter_lf)
    p("Trade 08 - Kitchen & Bar", 3, 900.0)
    q("Trade 08 - Kitchen & Bar", 4, grohe_faucet_ea)
    p("Trade 08 - Kitchen & Bar", 4, 900.0)
    p("Trade 08 - Kitchen & Bar", 5, 18000.0)  # appliances lot (qty=1)
    p("Trade 08 - Kitchen & Bar", 6, 65000.0)  # event bar lump

    # Trade 09 - Sanitary & Shower
    q("Trade 09 - Sanitary & Shower", 2, sanitary_tile_sqm)
    p("Trade 09 - Sanitary & Shower", 2, 250.0)
    q("Trade 09 - Sanitary & Shower", 3, sanitary_fixtures_ea)
    p("Trade 09 - Sanitary & Shower", 3, 1200.0)
    p("Trade 09 - Sanitary & Shower", 4, 15000.0)  # waterproofing/accessories lump

    # Trade 10 - Furniture (ex millwo
    q("Trade 10 - Furniture (ex millwo", 2, workstations_qty)
    p("Trade 10 - Furniture (ex millwo", 2, 250.0)
    q("Trade 10 - Furniture (ex millwo", 3, furniture_items_ea)
    p("Trade 10 - Furniture (ex millwo", 3, 85.0)
    q("Trade 10 - Furniture (ex millwo", 4, furniture_items_ea)
    p("Trade 10 - Furniture (ex millwo", 4, 60.0)
    q("Trade 10 - Furniture (ex millwo", 5, furniture_items_ea)
    p("Trade 10 - Furniture (ex millwo", 5, 20.0)

    # Trade 11 - Signage & Window Fil
    q("Trade 11 - Signage & Window Fil", 2, window_film_sqm)
    p("Trade 11 - Signage & Window Fil", 2, 85.0)
    q("Trade 11 - Signage & Window Fil", 3, logos_ea)
    p("Trade 11 - Signage & Window Fil", 3, 450.0)

    # Trade 12 - Technical & Misc
    p("Trade 12 - Technical & Misc", 2, 12000.0)
    p("Trade 12 - Technical & Misc", 3, 25000.0)
    q("Trade 12 - Technical & Misc", 4, racks_ea)
    p("Trade 12 - Technical & Misc", 4, 450.0)

    # Apply values + formulas
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Trade "):
            continue
        ws = wb[sheet_name]

        for (s, row), qty in quantities.items():
            if s == sheet_name:
                ws.cell(row=row, column=4).value = qty
        for (s, row), price in unit_prices.items():
            if s == sheet_name:
                ws.cell(row=row, column=5).value = price

        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value:
                ws.cell(r, 6).value = f"=D{r}*E{r}"

    # Recap formulas
    recap = wb["00 - Recap_EN"]
    lot_to_sheet = {
        1: "Trade 01 - Demolition",
        2: "Trade 02 - Partitions",
        3: "Trade 03 - Flooring",
        4: "Trade 04 - Ceilings",
        5: "Trade 05 - Custom Millwork",
        6: "Trade 06 - Paint & Walls",
        7: "Trade 07 - Electrical & Lightin",
        8: "Trade 08 - Kitchen & Bar",
        9: "Trade 09 - Sanitary & Shower",
        10: "Trade 10 - Furniture (ex millwo",
        11: "Trade 11 - Signage & Window Fil",
        12: "Trade 12 - Technical & Misc",
    }
    for recap_row in range(2, 14):
        lot = recap.cell(recap_row, 1).value
        if lot is None:
            continue
        try:
            lot_int = int(lot)
        except Exception:
            continue
        trade_sheet = lot_to_sheet.get(lot_int)
        if trade_sheet:
            recap.cell(recap_row, 6).value = f"=SUM('{trade_sheet}'!F:F)"

    recap["C16"].value = "=SUM(F2:F13)"
    recap["C17"].value = "=C16*0.05"
    recap["C18"].value = "=C16+C17"
    recap["C20"].value = "=C18"

    wb.save(output_path)

    assumptions_path = out_dir / "ASSUMPTIONS_AND_NOTES.txt"
    assumptions_path.write_text(
        \"\"\"SALOMON NYC – AI Estimate (Chapter cost baseline in client template)\n\nFiles used:\n- US_Summary description of works_Salomon NY_160126.pdf\n- SALOMON NYC ESQ - Ind. E - 15 janvier 2026.pdf\n- 26007 Survey.pdf\n\nMeasured quantities used:\n- Total floor area assumed for takeoff: 2,763 SF ≈ 256.6 m²\n- Kitchen floor area: 70.0 m² (ESQ)\n- 64.4 m² callout (ESQ) treated as Design Studio\n\nMajor assumptions to verify:\n- Solid partitions: 120 m²; Glazed partitions: 120 LF; Doors: 18 ea\n- Flooring allocations (m²): Terrazzo=100.0, Oak=25.0, Carpet(EGE)=64.4, Carpet(Interface)=remainder\n- Paint area (walls+ceilings): 2.5 × floor area\n- Window film area: 111.5 m² (approx. 120 LF × 10' glazing height)\n- Lighting counts: CLAREO 120 ea, TAL 40 ea, Decorative 25 ea\n- Sanitary tile: 55 m²; Sanitary fixtures: 20 ea\n- Furniture assembly: 30 workstations + 120 misc items\n- Haul-off: 20 tons\n\nPricing note:\n- Unit prices entered are first-pass AI budget assumptions for internal use; final bid requires drawing/spec takeoff + vendor quotes.\n\"\"\",\n        encoding="utf-8",\n    )

    print(f"Wrote filled template: {output_path}")
    print(f"Wrote assumptions: {assumptions_path}")


if __name__ == "__main__":
    main()


